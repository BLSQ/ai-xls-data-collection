"""OpenHEXA Pipeline: Excel Validation (AI-schema based).

Validates Excel files against schema_validation.json produced by the
aedes_schema_from_ai_template pipeline. Fully generalized — no hardcoded
domain terms (no "bailleur", "intitule_budgetaire", etc.).

Produces a JSON report with errors, warnings, and confirmations per sheet,
plus an extraction_guide.json for the ingest pipeline.
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openhexa.sdk import File, current_run, parameter, pipeline, workspace
from openpyxl.utils import column_index_from_string, get_column_letter

from validators import (
    find_data_rows,
    is_v2_schema,
    read_cell,
    read_cell_rc,
    text_matches,
    validate_data_rows,
    validate_headers,
    validate_labels,
    validate_sections_headers,
    validate_values,
)

SKIP_SHEETS = {"guide", "observations", "liste des catégories"}
SCHEMA_FILENAME = "schema_validation.json"


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------


@pipeline("xls-validation", timeout=3600)
@parameter(
    "excel_file",
    name="Excel File Path",
    type=File,
    required=True,
    help="Path to the Excel file to validate (relative to workspace files).",
)
def xls_validation(excel_file: File):
    """Validate an Excel file against the AI-generated schema."""
    current_run.log_info(f"Starting validation of: {excel_file}")

    excel_path = excel_file.path
    schema_file = Path(workspace.files_path) / SCHEMA_FILENAME

    schema = load_schema(schema_file)
    config = schema.get("validation_config", {}).get("label_comparison", {})

    current_run.log_info(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    report, extraction_guide = build_report(wb, schema, config)

    report_path = Path(workspace.files_path) / "validation_report.json"
    save_json(report, report_path, "validation report")

    guide_path = Path(workspace.files_path) / "extraction_guide.json"
    save_json(extraction_guide, guide_path, "extraction guide")

    log_summary(report)


def load_schema(schema_path: Path) -> dict:
    current_run.log_info(f"Loading schema from: {schema_path}")
    with schema_path.open(encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Report + extraction guide
# ---------------------------------------------------------------------------


def build_report(wb, schema: dict, config: dict) -> tuple[dict, dict]:
    """Validate all eligible sheets and return (report, extraction_guide)."""
    report = {
        "generated_at": datetime.now().isoformat(),
        "schema_version": schema.get("version", "unknown"),
        "sheets": {},
        "summary": {
            "total_sheets_validated": 0,
            "total_sheets_skipped": 0,
            "skipped_sheet_names": [],
            "total_errors": 0,
            "total_warnings": 0,
            "total_info": 0,
        },
    }

    extraction_guide = {
        "generated_at": datetime.now().isoformat(),
        "schema_version": schema.get("version", "unknown"),
        "sheets": {},
    }

    # Merge hardcoded skip list with schema-defined skip_sheets
    skip_sheets = set(SKIP_SHEETS)
    for s in schema.get("schema_meta", {}).get("skip_sheets", []):
        skip_sheets.add(s.strip().lower())

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in skip_sheets:
            report["summary"]["total_sheets_skipped"] += 1
            report["summary"]["skipped_sheet_names"].append(sheet_name)
            current_run.log_info(f"Skipping sheet: '{sheet_name}'")
            continue

        current_run.log_info(f"Validating sheet: '{sheet_name}'")
        ws = wb[sheet_name]

        # Fingerprint check: use the first label_fields module's title
        fingerprint = _get_fingerprint(schema)
        if fingerprint:
            fp_val = read_cell(ws, fingerprint["cell"])
            if fp_val is None or not text_matches(
                str(fp_val),
                fingerprint["expected"],
                fingerprint.get("known_variants", []),
                config,
            ):
                report["summary"]["total_sheets_skipped"] += 1
                report["summary"]["skipped_sheet_names"].append(sheet_name)
                current_run.log_info(
                    f"Sheet '{sheet_name}' does not match expected format. Skipping.",
                )
                continue

        sheet_issues, column_map, module_columns = validate_sheet(
            ws,
            schema,
            config,
        )
        data_rows = find_data_rows(ws, schema)

        # Detect grouped column structures for extraction guide
        grouped_info = _detect_grouped_columns(ws, schema)

        guide_sheet = build_extraction_guide(
            schema,
            column_map,
            module_columns,
            data_rows,
            grouped_info,
        )
        extraction_guide["sheets"][sheet_name] = guide_sheet

        error_count = sum(1 for i in sheet_issues if i["severity"] == "error")
        warning_count = sum(1 for i in sheet_issues if i["severity"] == "warning")
        info_count = sum(1 for i in sheet_issues if i["severity"] == "info")

        report["sheets"][sheet_name] = {
            "issues": sheet_issues,
            "counts": {
                "errors": error_count,
                "warnings": warning_count,
                "info": info_count,
                "total": len(sheet_issues),
            },
            "valid": error_count == 0,
        }

        report["summary"]["total_sheets_validated"] += 1
        report["summary"]["total_errors"] += error_count
        report["summary"]["total_warnings"] += warning_count
        report["summary"]["total_info"] += info_count

    return report, extraction_guide


def _get_fingerprint(schema: dict) -> dict | None:
    """Get the fingerprint for sheet matching.

    Supports both v1 (label_fields) and v2 (schema_meta.fingerprint) formats.
    """
    # v2: schema_meta.fingerprint
    fp = schema.get("schema_meta", {}).get("fingerprint")
    if fp and isinstance(fp, dict) and "cell" in fp:
        return fp

    # v1: first module title in label_fields
    label_fields = schema.get("label_fields", {})
    for mod_key, mod_def in label_fields.items():
        if mod_key.startswith("_"):
            continue
        if isinstance(mod_def, dict) and "title" in mod_def:
            title = mod_def["title"]
            if isinstance(title, dict) and "cell" in title:
                return title
    return None


def _detect_grouped_columns(ws, schema: dict) -> dict:
    """Detect grouped column structures from value_fields.

    Returns a dict mapping section_key to {columns, active_columns, anchor_row}.
    For v2 schemas (no value_fields), returns empty dict.
    """
    grouped_info = {}
    value_fields = schema.get("value_fields")
    if not value_fields:
        return grouped_info

    for section_key, section in value_fields.items():
        if not isinstance(section, dict):
            continue
        for field_key, field_def in section.items():
            if not isinstance(field_def, dict):
                continue
            if "columns" not in field_def or "rows" not in field_def:
                continue

            # This is a grouped entry
            sec_key = (
                field_key.removesuffix("_values") if field_key.endswith("_values") else field_key
            )
            columns = field_def["columns"]
            rows_def = field_def.get("rows", {})

            # Find anchor row
            anchor_row = None
            for rk, rd in rows_def.items():
                if rd.get("required_rule") == "anchor":
                    anchor_row = rd.get("row")
                    break

            active_cols = []
            if anchor_row:
                for col in columns:
                    val = read_cell_rc(ws, anchor_row, col)
                    if val is not None and str(val).strip():
                        active_cols.append(col)

            grouped_info[sec_key] = {
                "columns": columns,
                "active_columns": active_cols,
                "anchor_row": anchor_row,
            }

    return grouped_info


def validate_sheet(
    ws,
    schema: dict,
    config: dict,
) -> tuple[list[dict], dict, dict]:
    """Run all validation passes. Returns (issues, column_map, module_columns).

    Dispatches to v2 (sections-based) or v1 (flat-keys) validation paths.
    """
    issues = []

    if is_v2_schema(schema):
        # v2: sections-based schema — no label_fields or value_fields
        header_issues, column_map = validate_sections_headers(ws, schema, config)
        issues.extend(header_issues)
        issues.extend(validate_data_rows(ws, schema, column_map, {}))
        return issues, column_map, {}

    # v1: flat-keys schema
    issues.extend(validate_labels(ws, schema, config))
    issues.extend(validate_values(ws, schema, config))

    header_issues, column_map, module_columns = validate_headers(
        ws,
        schema,
        config,
    )
    issues.extend(header_issues)

    issues.extend(validate_data_rows(ws, schema, column_map, module_columns))

    return issues, column_map, module_columns


# ---------------------------------------------------------------------------
# Extraction guide builder
# ---------------------------------------------------------------------------


def build_extraction_guide(
    schema: dict,
    column_map: dict,
    module_columns: dict,
    data_rows: list[int],
    grouped_info: dict,
) -> dict:
    """Build an extraction guide for a single sheet.

    Supports both v1 (header_modules/table_definitions) and v2 (sections) schemas.
    """
    guide = {
        "metadata": _build_metadata_guide(schema, grouped_info),
        "data": {
            "start_row": min(data_rows) if data_rows else 2,
            "end_row": max(data_rows) if data_rows else 2,
            "row_count": len(data_rows),
            "columns": {},
        },
    }

    if is_v2_schema(schema):
        # v2: build columns from sections
        for section in schema.get("sections", []):
            if section.get("layout") != "records":
                continue
            for col_def in section.get("headers", {}).get("columns", []):
                _walk_v2_column_entry(guide["data"]["columns"], col_def)
    else:
        # v1: build columns from header_modules
        header_modules = schema.get("header_modules", {})
        for mod_key, mod_def in header_modules.items():
            if mod_key.startswith("_") or not isinstance(mod_def, dict):
                continue
            if "columns" not in mod_def:
                continue
            col_cursor = column_index_from_string(mod_def.get("start_column", "B"))
            for col_entry in mod_def["columns"]:
                col_cursor = _walk_column_entry(
                    guide["data"]["columns"],
                    col_entry,
                    mod_key,
                    "default",
                    column_map,
                    col_cursor,
                )

    return guide


def _walk_v2_column_entry(columns_dict: dict, col_def: dict):
    """Walk a v2 column entry and populate columns_dict."""
    col_type = col_def.get("type", "single")

    if col_type == "single":
        canonical = col_def.get("canonical_name", "")
        col_letter = col_def.get("column", "")
        if canonical and col_letter:
            columns_dict[canonical] = {
                "col_letter": col_letter,
                "header": col_def.get("expected", ""),
                "value_type": col_def.get("value_type", "string"),
                "required": col_def.get("required", False),
            }
    elif col_type == "group":
        parent = col_def.get("parent", {})
        group_name = parent.get("expected", "")
        for sub in col_def.get("sub_columns", []):
            canonical = sub.get("canonical_name", "")
            sub_col = sub.get("column", "")
            parent_canonical = parent.get("canonical_name", "")
            full_canonical = f"{parent_canonical}__{canonical}" if parent_canonical else canonical
            if full_canonical and sub_col:
                columns_dict[full_canonical] = {
                    "col_letter": sub_col,
                    "header": sub.get("expected", ""),
                    "value_type": sub.get("value_type", "string"),
                    "group": group_name,
                }


def _build_metadata_guide(schema: dict, grouped_info: dict) -> dict:
    """Build the metadata section of the extraction guide.

    Reads from table_definitions.metadata generically.
    """
    td = schema.get("table_definitions", {}).get("metadata", {})
    fields = []
    for f in td.get("fields", []):
        fields.append(
            {
                "canonical_name": f["canonical_name"],
                "value_cell": f["value_cell"],
                "type": f["type"],
            }
        )

    # Build grouped sections generically
    groups = {}
    for sec_key, info in grouped_info.items():
        # Look for fields_per_column in table_definitions.metadata
        fields_key = f"{sec_key}_fields"
        td_grouped = td.get(fields_key, {})
        groups[sec_key] = {
            "columns": info["columns"],
            "active_columns": info["active_columns"],
            "fields_per_column": td_grouped.get("fields_per_column", []),
        }

    result = {"fields": fields}
    if groups:
        result["grouped_sections"] = groups

    return result


def _walk_column_entry(
    columns_dict: dict,
    col_entry: dict,
    module: str,
    category: str,
    column_map: dict,
    col_cursor: int,
) -> int:
    """Walk a column entry and populate columns_dict. Returns next col_cursor."""
    pos = col_entry.get("position", "fixed")

    if pos == "fixed":
        canonical = col_entry.get("canonical_name", "")
        col_letter = col_entry.get("column") or get_column_letter(col_cursor)
        if canonical:
            columns_dict[canonical] = {
                "col_letter": col_letter,
                "module": module,
                "category": category,
                "header": col_entry.get("expected", ""),
                "value_type": col_entry.get("value_type", "string"),
            }
        if col_entry.get("column"):
            return column_index_from_string(col_entry["column"]) + 1
        return col_cursor + 1

    elif pos == "group":
        parent = col_entry.get("parent", {})
        group_category = parent.get("expected", category)
        cursor = col_cursor
        for sub in col_entry.get("sub_columns", []):
            cursor = _walk_column_entry(
                columns_dict,
                sub,
                module,
                group_category,
                column_map,
                cursor,
            )
        return cursor

    elif pos == "dynamic":
        prefix = col_entry.get("canonical_prefix", "")
        vtype = col_entry.get("value_type", "string")
        count = 0
        for canonical, col_letter in column_map.items():
            if canonical.startswith(prefix + "_"):
                header = canonical[len(prefix) + 1 :].replace("_", " ").title()
                columns_dict[canonical] = {
                    "col_letter": col_letter,
                    "module": module,
                    "category": category,
                    "header": header,
                    "value_type": vtype,
                }
                count += 1
        return col_cursor + count

    return col_cursor


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------


def save_json(data: dict, output_path: Path, label: str):
    current_run.log_info(f"Saving {label} to: {output_path}")
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    current_run.add_file_output(str(output_path))


def log_summary(report: dict):
    s = report["summary"]
    current_run.log_info(
        f"Validation complete — "
        f"{s['total_sheets_validated']} sheets validated, "
        f"{s['total_sheets_skipped']} skipped"
    )
    if s["total_errors"] > 0:
        current_run.log_warning(f"Found {s['total_errors']} errors across all sheets")
    if s["total_warnings"] > 0:
        current_run.log_warning(f"Found {s['total_warnings']} warnings requiring confirmation")
    if s["total_errors"] == 0 and s["total_warnings"] == 0:
        current_run.log_info("All sheets passed validation with no errors or warnings.")


if __name__ == "__main__":
    xls_validation()
