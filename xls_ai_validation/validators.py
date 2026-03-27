"""Validation helpers for AI-schema-based Excel validation.

Generalized version — no hardcoded domain terms. Detects grouped entries
structurally (by presence of 'columns'+'rows' keys) rather than by name.

Pure functions with no OpenHEXA SDK dependency — can be tested standalone.
"""

import re
import unicodedata
from datetime import datetime

from openpyxl.utils import column_index_from_string, get_column_letter


# ---------------------------------------------------------------------------
# Text normalization helpers
# ---------------------------------------------------------------------------


def normalize_text(text: str, config: dict) -> str:
    """Normalize a cell value for comparison according to validation_config."""
    if text is None:
        return ""
    s = str(text)
    if config.get("strip_whitespace", True):
        s = s.strip()
        s = re.sub(r"\s+", " ", s)
    if config.get("strip_trailing_colon", True):
        s = s.rstrip(":")
    if not config.get("case_sensitive"):
        s = s.lower()
    if config.get("normalize_accents", True):
        s = _remove_accents(s)
    return s


def _remove_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


_LIGATURES = str.maketrans({"œ": "oe", "Œ": "OE", "æ": "ae", "Æ": "AE"})


def to_canonical(text: str) -> str:
    """Deterministic canonical name from free-text (matches schema generator)."""
    if not text:
        return "unnamed"
    text = str(text).strip()
    nfkd = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in nfkd if not unicodedata.combining(c))
    text = text.translate(_LIGATURES)
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_").lower()
    return text or "unnamed"


def text_matches(actual: str, expected: str, variants: list[str], config: dict) -> str:
    """Check if actual matches expected or any variant.

    Returns "exact", "variant", or "" (no match).
    """
    norm_actual = normalize_text(actual, config)
    if norm_actual == normalize_text(expected, config):
        return "exact"
    for v in variants:
        if norm_actual == normalize_text(v, config):
            return "variant"
    return ""


# ---------------------------------------------------------------------------
# JsonLogic evaluation
# ---------------------------------------------------------------------------


def evaluate_json_logic(rule, data):
    """Evaluate a JsonLogic rule dict against a data context dict.

    Supports standard operators (==, !=, >, >=, <, <=, +, -, *, /, !, if,
    and, or) plus var for resolving column values by canonical name.
    """
    if not isinstance(rule, dict):
        return rule

    op = next(iter(rule))
    raw_args = rule[op]
    if not isinstance(raw_args, list):
        raw_args = [raw_args]

    if op == "var":
        path = raw_args[0] if raw_args else ""
        default = raw_args[1] if len(raw_args) > 1 else None
        result = _jl_var(data, path)
        return result if result is not None else default

    if op in ("and", "or"):
        return _jl_short_circuit(op, raw_args, data)

    args = [evaluate_json_logic(a, data) for a in raw_args]

    if op == "==":
        return args[0] == args[1] if len(args) >= 2 else False
    if op == "!=":
        return args[0] != args[1] if len(args) >= 2 else True
    if op == ">":
        return float(args[0]) > float(args[1])
    if op == ">=":
        return float(args[0]) >= float(args[1])
    if op == "<":
        return float(args[0]) < float(args[1])
    if op == "<=":
        return float(args[0]) <= float(args[1])
    if op == "+":
        return sum(float(a) for a in args if a is not None)
    if op == "-":
        return -float(args[0]) if len(args) == 1 else float(args[0]) - float(args[1])
    if op == "*":
        result = 1.0
        for a in args:
            result *= float(a)
        return result
    if op == "/":
        return float(args[0]) / float(args[1])
    if op == "!":
        return not args[0]
    if op == "if":
        i = 0
        while i < len(args) - 1:
            if args[i]:
                return args[i + 1]
            i += 2
        return args[i] if i < len(args) else None

    raise ValueError(f"Unknown JsonLogic operator: {op}")


def _jl_var(data, path):
    if not path:
        return data
    result = data
    for k in str(path).split("."):
        if isinstance(result, dict):
            result = result.get(k)
        else:
            return None
    return result


def _jl_short_circuit(op, raw_args, data):
    if op == "and":
        result = True
        for arg in raw_args:
            result = evaluate_json_logic(arg, data)
            if not result:
                return result
        return result
    else:
        result = False
        for arg in raw_args:
            result = evaluate_json_logic(arg, data)
            if result:
                return result
        return result


# ---------------------------------------------------------------------------
# Cell reading helpers
# ---------------------------------------------------------------------------


def read_cell(ws, cell_ref: str):
    """Read a cell value from a worksheet by its A1 reference."""
    return ws[cell_ref].value


def read_cell_rc(ws, row: int, col_letter: str):
    """Read a cell value by row number and column letter."""
    col_idx = column_index_from_string(col_letter)
    return ws.cell(row=row, column=col_idx).value


def cell_ref(col_letter: str, row: int) -> str:
    return f"{col_letter}{row}"


# ---------------------------------------------------------------------------
# Issue builder
# ---------------------------------------------------------------------------


def make_issue(
    severity: str,
    cell: str,
    message: str,
    field_ref: str = "",
    group: str = "",
) -> dict:
    return {
        "severity": severity,
        "cell": cell,
        "message": message,
        "field_ref": field_ref,
        "group": group,
    }


# ---------------------------------------------------------------------------
# Label validation
# ---------------------------------------------------------------------------


def validate_labels(ws, schema: dict, config: dict) -> list[dict]:
    """Validate that all label_fields are present at their expected cells.

    Safely returns [] if the schema has no label_fields (v2 sections-based schemas).
    """
    issues = []
    label_fields = schema.get("label_fields")
    if not label_fields:
        return issues

    for module_key, module_def in label_fields.items():
        if module_key.startswith("_"):
            continue

        # Entries with "items" list (e.g., *_columns entries)
        if "items" in module_def:
            for item in module_def["items"]:
                _check_label(ws, item, config, issues, module_key)
            continue

        for field_key, field_def in module_def.items():
            if field_key.startswith("_"):
                continue
            if not isinstance(field_def, dict) or "cell" not in field_def:
                continue
            _check_label(ws, field_def, config, issues, f"{module_key}.{field_key}")

    return issues


def _check_label(ws, field_def: dict, config: dict, issues: list, field_ref: str):
    cell = field_def["cell"]
    if not cell:
        return
    expected = field_def["expected"]
    variants = field_def.get("known_variants", [])
    actual = read_cell(ws, cell)

    if actual is not None:
        match = text_matches(str(actual), expected, variants, config)
        if match:
            return
        issues.append(
            make_issue(
                "warning",
                cell,
                f"Le label en {cell} ('{actual}') ne correspond pas "
                f"exactement à '{expected}'. Vérifiez la cellule.",
                field_ref,
                "label",
            )
        )
        return

    # Cell is empty — scan ±3 columns on the same row
    m = re.match(r"^([A-Z]+)(\d+)$", cell)
    if m:
        col_idx = column_index_from_string(m.group(1))
        row = int(m.group(2))
        for offset in (-1, 1, -2, 2, -3, 3):
            try_idx = col_idx + offset
            if try_idx < 1:
                continue
            try_col = get_column_letter(try_idx)
            try_val = read_cell_rc(ws, row, try_col)
            if try_val is None:
                continue
            match = text_matches(str(try_val), expected, variants, config)
            if match:
                issues.append(
                    make_issue(
                        "info",
                        f"{try_col}{row}",
                        f"Label '{expected}' trouvé en {try_col}{row} "
                        f"au lieu de {cell} (décalage de {offset} colonne(s)).",
                        field_ref,
                        "label",
                    )
                )
                return

    issues.append(
        make_issue(
            "error",
            cell,
            f"Label attendu manquant en {cell}. Attendu : '{expected}'.",
            field_ref,
            "label",
        )
    )


# ---------------------------------------------------------------------------
# Value validation
# ---------------------------------------------------------------------------


def validate_values(ws, schema: dict, config: dict) -> list[dict]:
    """Validate value_fields: null checks, type checks, regex/enum validation.

    Generically detects grouped entries (those with 'columns' and 'rows' keys)
    vs single-cell entries (those with a 'cell' key).
    """
    issues = []
    value_fields = schema.get("value_fields")
    if not value_fields:
        return issues

    for section_key, section in value_fields.items():
        if section_key.startswith("_"):
            continue
        if not isinstance(section, dict):
            continue

        for field_key, field_def in section.items():
            if field_key.startswith("_"):
                continue
            if not isinstance(field_def, dict):
                continue

            # Detect grouped entry: has 'columns' and 'rows'
            if "columns" in field_def and "rows" in field_def:
                issues.extend(_validate_grouped_values(ws, field_def, config))
            elif "cell" in field_def:
                issues.extend(_validate_single_value(ws, field_key, field_def, config))

    return issues


def _validate_single_value(
    ws,
    field_key: str,
    field_def: dict,
    config: dict,
) -> list[dict]:
    """Validate a single-cell value entry."""
    issues = []
    cell = field_def["cell"]
    value = read_cell(ws, cell)
    label_display = field_def.get("label_display", field_key)
    field_ref = field_def.get("label_ref", field_key)
    required = field_def.get("required", False)
    field_type = field_def.get("type", "string")
    validation = field_def.get("validation", {})
    message = field_def.get("message", "")

    # Null check
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            issues.append(make_issue("error", cell, message, field_ref, "value"))
        else:
            issues.append(make_issue("info", cell, message, field_ref, "value"))
        return issues

    # Type check
    example = field_def.get("example", "")
    example_hint = f" Exemple attendu : '{example}'." if example else ""

    if field_type == "number":
        try:
            num_val = float(value)
            if "min" in validation and num_val < validation["min"]:
                issues.append(make_issue("error", cell, message, field_ref, "value"))
        except (ValueError, TypeError):
            issues.append(
                make_issue(
                    "error",
                    cell,
                    f"Le champ '{label_display}' ({cell}) doit être "
                    f"un nombre. Valeur actuelle : '{value}'.{example_hint}",
                    field_ref,
                    "value",
                )
            )

    elif field_type == "date":
        if not isinstance(value, datetime) and not _is_date_string(value):
            msg = (
                (
                    f"Le champ '{label_display}' ({cell}) doit contenir une "
                    f"date valide. Valeur actuelle : '{value}'.{example_hint}"
                )
                if example_hint
                else message
            )
            issues.append(make_issue("error", cell, msg, field_ref, "value"))

    elif field_type == "string":
        str_val = str(value).strip()
        if "min_length" in validation and len(str_val) < validation["min_length"]:
            issues.append(make_issue("error", cell, message, field_ref, "value"))
        if "regex" in validation and not re.match(validation["regex"], str_val):
            msg = (
                (
                    f"Le champ '{label_display}' ({cell}) ne correspond pas "
                    f"au format attendu. Valeur actuelle : '{str_val}'.{example_hint}"
                )
                if example_hint
                else message
            )
            issues.append(make_issue("error", cell, msg, field_ref, "value"))
        if "enum" in validation and str_val not in validation["enum"]:
            issues.append(
                make_issue(
                    "error",
                    cell,
                    f"Le champ '{label_display}' ({cell}) contient "
                    f"'{str_val}'. Valeurs acceptées : "
                    f"{validation['enum']}.{example_hint}",
                    field_ref,
                    "value",
                )
            )

    # Cell-level json_logic rules
    jl_def = validation.get("json_logic")
    if jl_def:
        jl_entries = jl_def if isinstance(jl_def, list) else [jl_def]
        jl_combine = validation.get("json_logic_op", "and")
        cell_data = {"value": value}
        try:
            jl_results = [evaluate_json_logic(e.get("rule", e), cell_data) for e in jl_entries]
            jl_result = any(jl_results) if jl_combine == "or" else all(jl_results)
            if not jl_result:
                failed_descs = [
                    e.get("description", "")
                    for e, r in zip(jl_entries, jl_results)
                    if not r and e.get("description")
                ]
                detail = "; ".join(failed_descs) if failed_descs else "règle non satisfaite"
                issues.append(
                    make_issue(
                        "error",
                        cell,
                        f"Le champ '{label_display}' ({cell}): {detail}",
                        field_ref,
                        "value",
                    )
                )
        except Exception as exc:
            issues.append(
                make_issue(
                    "warning",
                    cell,
                    f"Erreur JsonLogic pour '{label_display}' ({cell}): {exc}",
                    field_ref,
                    "value",
                )
            )

    return issues


def _is_date_string(value) -> bool:
    if isinstance(value, datetime):
        return True
    try:
        datetime.fromisoformat(str(value))
        return True
    except (ValueError, TypeError):
        return False


def _validate_grouped_values(
    ws,
    grouped_def: dict,
    config: dict,
) -> list[dict]:
    """Validate a grouped entry (columns × rows structure).

    Generalized version of the old _validate_bailleur_values — works for
    any repeating column group (donors, partners, regions, etc.).
    """
    issues = []
    columns = grouped_def["columns"]
    rows_def = grouped_def.get("rows", {})

    # Find the anchor row: the one with required_rule == "anchor"
    anchor_row_key = None
    for rk, rd in rows_def.items():
        if rd.get("required_rule") == "anchor":
            anchor_row_key = rk
            break

    if anchor_row_key is None:
        return issues  # No anchor rule — skip

    anchor_row_def = rows_def[anchor_row_key]
    anchor_row = anchor_row_def["row"]

    # Determine which columns are "active" (have a value in the anchor row)
    active_cols = []
    for col in columns:
        val = read_cell_rc(ws, anchor_row, col)
        if val is not None and str(val).strip():
            active_cols.append(col)

    # anchor: at least one column must have a value
    if not active_cols:
        issues.append(
            make_issue(
                "error",
                cell_ref(columns[0], anchor_row),
                anchor_row_def.get("message", ""),
                anchor_row_def.get("label_ref", ""),
                "value",
            )
        )

    # For each other row: validate active columns
    for row_key, row_def in rows_def.items():
        if row_key == anchor_row_key:
            continue

        rule = row_def.get("required_rule", "")
        if rule != "if_active":
            continue

        row_num = row_def["row"]
        for col in active_cols:
            val = read_cell_rc(ws, row_num, col)
            if val is None or (isinstance(val, str) and not val.strip()):
                issues.append(
                    make_issue(
                        "error",
                        cell_ref(col, row_num),
                        f"{row_def.get('message', '')} (colonne {col})",
                        row_def.get("label_ref", ""),
                        "value",
                    )
                )
            elif row_def.get("type") == "number":
                try:
                    num_val = float(val)
                    min_val = row_def.get("validation", {}).get("min")
                    if min_val is not None and num_val < min_val:
                        issues.append(
                            make_issue(
                                "error",
                                cell_ref(col, row_num),
                                f"{row_def.get('message', '')} (colonne {col}) — valeur négative.",
                                row_def.get("label_ref", ""),
                                "value",
                            )
                        )
                except (ValueError, TypeError):
                    example = row_def.get("example", "")
                    hint = f" Exemple attendu : '{example}'." if example else ""
                    issues.append(
                        make_issue(
                            "error",
                            cell_ref(col, row_num),
                            f"{row_def.get('label_display', '')} ({col}{row_num}) "
                            f"doit être un nombre.{hint}",
                            row_def.get("label_ref", ""),
                            "value",
                        )
                    )

    return issues


# ---------------------------------------------------------------------------
# Header validation
# ---------------------------------------------------------------------------


def validate_headers(
    ws,
    schema: dict,
    config: dict,
) -> tuple[list[dict], dict, dict]:
    """Validate header_modules and build canonical_name -> column_letter mapping.

    Handles both old format (position: fixed/group/dynamic) and new format
    (type: single/group with fixed: true/false sub_columns).

    Returns (issues, column_map, module_col_letters).
    """
    issues = []
    column_map = {}
    module_col_letters = {}
    header_modules = schema.get("header_modules", {})
    template = header_modules.get("mismatch_message_template", "")

    for mod_key, mod_def in header_modules.items():
        if mod_key.startswith("_"):
            continue
        if not isinstance(mod_def, dict) or "columns" not in mod_def:
            continue

        mod_letters = []

        # Validate module title
        title_def = mod_def.get("title")
        if title_def:
            title_cell = title_def["cell"]
            title_val = read_cell(ws, title_cell)
            if title_val is not None:
                match = text_matches(str(title_val), title_def["expected"], [], config)
                if not match:
                    issues.append(
                        make_issue(
                            "warning",
                            title_cell,
                            f"Titre du {mod_key} en {title_cell} ('{title_val}') "
                            f"ne correspond pas à '{title_def['expected']}'.",
                            mod_key,
                            "header",
                        )
                    )

        col_cursor = column_index_from_string(mod_def.get("start_column", "B"))

        for col_entry in mod_def["columns"]:
            position = col_entry.get("position", "fixed")

            if position == "fixed":
                _validate_fixed_header(
                    ws,
                    col_entry,
                    config,
                    template,
                    issues,
                    column_map,
                )
                col_letter = col_entry.get("column")
                if col_letter:
                    mod_letters.append(col_letter)
                    col_cursor = column_index_from_string(col_letter) + 1

            elif position == "group":
                _, col_cursor, group_letters = _validate_group_header(
                    ws,
                    col_entry,
                    col_cursor,
                    config,
                    template,
                    issues,
                    column_map,
                    mod_key,
                )
                mod_letters.extend(group_letters)

            elif position == "dynamic":
                dyn_canonicals, col_cursor = _validate_dynamic_headers(
                    ws,
                    col_entry,
                    col_cursor,
                    issues,
                    column_map,
                )
                for cn in dyn_canonicals:
                    cl = column_map.get(cn)
                    if cl:
                        mod_letters.append(cl)

        module_col_letters[mod_key] = mod_letters

    return issues, column_map, module_col_letters


def _validate_fixed_header(
    ws,
    col_def: dict,
    config: dict,
    template: str,
    issues: list,
    column_map: dict,
) -> list[str]:
    """Validate a single fixed-position header."""
    col_letter = col_def["column"]
    row = col_def["row"]
    expected = col_def["expected"]
    canonical = col_def["canonical_name"]
    variants = col_def.get("known_variants", [])

    actual = read_cell_rc(ws, row, col_letter)
    column_map[canonical] = col_letter

    if actual is None:
        issues.append(
            make_issue(
                "error",
                cell_ref(col_letter, row),
                f"En-tête manquant en {col_letter}{row}. Attendu : '{expected}' (→ {canonical}).",
                canonical,
                "header",
            )
        )
        return [canonical]

    actual_str = str(actual).strip()
    match = text_matches(actual_str, expected, variants, config)

    if match == "variant":
        issues.append(
            make_issue(
                "info",
                cell_ref(col_letter, row),
                f"En-tête en {col_letter}{row} ('{actual_str}') "
                f"reconnu comme variante de '{expected}'.",
                canonical,
                "header",
            )
        )
    elif not match:
        msg = (
            template.format(
                column=col_letter,
                actual_value=actual_str,
                expected=expected,
                canonical_name=canonical,
            )
            if template
            else (
                f"En-tête en {col_letter}{row} ('{actual_str}') ne correspond pas à '{expected}'."
            )
        )
        issues.append(make_issue("warning", cell_ref(col_letter, row), msg, canonical, "header"))

    return [canonical]


def _validate_group_header(
    ws,
    group_def: dict,
    col_cursor: int,
    config: dict,
    template: str,
    issues: list,
    column_map: dict,
    mod_key: str,
) -> tuple[list[str], int, list[str]]:
    """Validate a group (parent + sub_columns)."""
    parent = group_def["parent"]
    parent_row = parent["row"]
    parent_expected = parent["expected"]
    parent_variants = parent.get("known_variants", [])
    canonicals = []
    col_letters = []

    parent_col = get_column_letter(col_cursor)
    parent_val = read_cell_rc(ws, parent_row, parent_col)

    if parent_val is not None:
        match = text_matches(str(parent_val), parent_expected, parent_variants, config)
        if match == "variant":
            issues.append(
                make_issue(
                    "info",
                    cell_ref(parent_col, parent_row),
                    f"En-tête parent en {parent_col}{parent_row} ('{parent_val}') "
                    f"reconnu comme variante de '{parent_expected}'.",
                    mod_key,
                    "header",
                )
            )
        elif not match:
            issues.append(
                make_issue(
                    "warning",
                    cell_ref(parent_col, parent_row),
                    f"En-tête parent en {parent_col}{parent_row} ('{parent_val}') "
                    f"ne correspond pas à '{parent_expected}'.",
                    mod_key,
                    "header",
                )
            )

    cursor = col_cursor
    for sub in group_def.get("sub_columns", []):
        sub_pos = sub.get("position", "fixed")

        if sub_pos == "fixed":
            # Prefer producer's explicit column; fall back to cursor
            sub_col = sub.get("column") or get_column_letter(cursor)
            sub_with_col = {**sub, "column": sub_col}
            sub_canonicals = _validate_fixed_header(
                ws,
                sub_with_col,
                config,
                template,
                issues,
                column_map,
            )
            canonicals.extend(sub_canonicals)
            col_letters.append(sub_col)
            cursor = column_index_from_string(sub_col) + 1

        elif sub_pos == "dynamic":
            dyn_canonicals, cursor = _validate_dynamic_headers(
                ws,
                sub,
                cursor,
                issues,
                column_map,
            )
            canonicals.extend(dyn_canonicals)
            for cn in dyn_canonicals:
                cl = column_map.get(cn)
                if cl:
                    col_letters.append(cl)

    return canonicals, cursor, col_letters


def _validate_dynamic_headers(
    ws,
    dyn_def: dict,
    col_cursor: int,
    issues: list,
    column_map: dict,
) -> tuple[list[str], int]:
    """Scan consecutive cells matching a regex pattern."""
    row = dyn_def["row"]
    pattern = dyn_def.get("pattern", "")
    prefix = dyn_def.get("canonical_prefix", "dynamic")
    min_count = dyn_def.get("min_count", 0)
    canonicals = []
    cursor = col_cursor

    while cursor <= 16384:
        col_letter = get_column_letter(cursor)
        val = read_cell_rc(ws, row, col_letter)

        if val is None:
            break

        val_str = str(val).strip()
        if not re.match(pattern, val_str):
            break

        normalized = to_canonical(val_str)
        canonical = f"{prefix}_{normalized}"
        column_map[canonical] = col_letter
        canonicals.append(canonical)
        cursor += 1

    if len(canonicals) < min_count:
        col_letter = get_column_letter(col_cursor)
        issues.append(
            make_issue(
                "error",
                cell_ref(col_letter, row),
                f"Au moins {min_count} colonne(s) dynamique(s) attendue(s) "
                f"(pattern: {pattern}) à partir de {col_letter}{row}, "
                f"mais {len(canonicals)} trouvée(s).",
                prefix,
                "header",
            )
        )

    return canonicals, cursor


# ---------------------------------------------------------------------------
# Data row enum validation helper
# ---------------------------------------------------------------------------


def _check_column_enum(ws, col_entry, column_map, data_rows, issues, mod_key):
    """Check enum constraints on data row cells for a header column."""
    pos = col_entry.get("position", "fixed")

    if pos == "fixed":
        validation = col_entry.get("validation", {})
        enum_values = validation.get("enum")
        cond_enum = validation.get("conditional_enum")

        if not enum_values and not cond_enum:
            return
        col_letter = col_entry.get("column")
        if not col_letter:
            canonical = col_entry.get("canonical_name", "")
            col_letter = column_map.get(canonical)
        if not col_letter:
            return
        header_name = col_entry.get("expected", col_entry.get("canonical_name", ""))

        if cond_enum:
            # Conditional enum: allowed values depend on another column's value
            parent_col = cond_enum.get("depends_on_column")
            values_by_parent = cond_enum.get("values_by_parent", {})
            for row in data_rows:
                val = read_cell_rc(ws, row, col_letter)
                if val is None or (isinstance(val, str) and not val.strip()):
                    continue
                str_val = str(val).strip()
                parent_val = read_cell_rc(ws, row, parent_col)
                parent_str = str(parent_val).strip() if parent_val else ""
                allowed = values_by_parent.get(parent_str)
                if allowed is None:
                    # Unknown parent value — warn but don't block
                    issues.append(
                        make_issue(
                            "warning",
                            cell_ref(col_letter, row),
                            f"Ligne {row}, colonne '{header_name}' ({col_letter}): "
                            f"la valeur parente '{parent_str}' (col {parent_col}) "
                            f"n'a pas de liste de valeurs conditionnelles connue.",
                            mod_key,
                            "data",
                        )
                    )
                elif str_val not in allowed:
                    issues.append(
                        make_issue(
                            "error",
                            cell_ref(col_letter, row),
                            f"Ligne {row}, colonne '{header_name}' ({col_letter}): "
                            f"'{str_val}' n'est pas autorisé quand "
                            f"'{parent_str}' est sélectionné (col {parent_col}). "
                            f"Valeurs acceptées : {allowed}.",
                            mod_key,
                            "data",
                        )
                    )
        else:
            # Simple flat enum
            for row in data_rows:
                val = read_cell_rc(ws, row, col_letter)
                if val is None or (isinstance(val, str) and not val.strip()):
                    continue
                str_val = str(val).strip()
                if str_val not in enum_values:
                    issues.append(
                        make_issue(
                            "error",
                            cell_ref(col_letter, row),
                            f"Ligne {row}, colonne '{header_name}' ({col_letter}): "
                            f"'{str_val}' n'est pas une valeur autorisée. "
                            f"Valeurs acceptées : {enum_values}.",
                            mod_key,
                            "data",
                        )
                    )

    elif pos == "group":
        for sub in col_entry.get("sub_columns", []):
            _check_column_enum(ws, sub, column_map, data_rows, issues, mod_key)


# ---------------------------------------------------------------------------
# Data row validation
# ---------------------------------------------------------------------------


def validate_data_rows(
    ws,
    schema: dict,
    column_map: dict,
    module_col_letters: dict,
) -> list[dict]:
    """Validate data rows using records_data rules or v2 sections.

    Validates primary key presence, enum constraints, column-level JsonLogic
    rules, and module-level percentage rules.
    For v2 schemas (sections-based), validates required columns are non-empty.
    """
    issues = []

    # v2 sections-based schema
    if "sections" in schema and "table_definitions" not in schema:
        return _validate_data_rows_v2(ws, schema, column_map)

    records_def = schema.get("table_definitions", {}).get("records_data")
    if not records_def:
        return issues
    row_val = records_def.get("row_validation", {})
    start_row = records_def.get("data_row_start", 17)

    # Determine data rows
    data_rows = []
    for row in range(start_row, ws.max_row + 1):
        if _is_row_empty(ws, row):
            break
        data_rows.append(row)

    if not data_rows:
        issues.append(
            make_issue(
                "warning",
                f"B{start_row}",
                f"Aucune ligne de données trouvée à partir de la ligne {start_row}.",
                "records_data",
                "data",
            )
        )
        return issues

    # Validate primary key column is not null (generic — reads key from schema)
    for pk_key, pk_rule in row_val.items():
        if pk_key.startswith("_"):
            continue
        if not isinstance(pk_rule, dict):
            continue
        if pk_rule.get("required") is not True:
            continue

        pk_col_letter = column_map.get(pk_key)
        if not pk_col_letter:
            issues.append(
                make_issue(
                    "warning",
                    f"B{start_row}",
                    f"Clé primaire '{pk_key}' introuvable dans les en-têtes — "
                    f"validation des lignes de données impossible pour cette clé.",
                    pk_key,
                    "data",
                )
            )
            continue
        pk_message = pk_rule.get("message", f"Valeur manquante pour '{pk_key}' en ligne {{row}}.")
        for row in data_rows:
            val = read_cell_rc(ws, row, pk_col_letter)
            if val is None or (isinstance(val, str) and not val.strip()):
                issues.append(
                    make_issue(
                        "error",
                        cell_ref(pk_col_letter, row),
                        pk_message,
                        pk_key,
                        "data",
                    )
                )

    # Validate enum constraints on data row cells
    for mod_key, mod_def in schema.get("header_modules", {}).items():
        if mod_key.startswith("_") or not isinstance(mod_def, dict):
            continue
        for col_entry in mod_def.get("columns", []):
            _check_column_enum(ws, col_entry, column_map, data_rows, issues, mod_key)

    # Validate column-level json_logic rules (stored on individual column entries)
    for mod_key, mod_def in schema.get("header_modules", {}).items():
        if mod_key.startswith("_") or not isinstance(mod_def, dict):
            continue
        for col_entry in _iter_all_columns(mod_def):
            col_validation = col_entry.get("validation") or {}
            json_logic_def = col_validation.get("json_logic")
            if not json_logic_def:
                continue
            canonical = col_entry.get("canonical_name", "")
            # Prefer the explicit column letter stored in the schema entry over
            # column_map, which collapses duplicate canonical names (e.g. two
            # "Total ligne" columns in the same module both named "total_ligne").
            col_letter = col_entry.get("column") or column_map.get(canonical)
            if not col_letter:
                continue
            entries = json_logic_def if isinstance(json_logic_def, list) else [json_logic_def]
            combine_op = col_validation.get("json_logic_op", "and")
            for row in data_rows:
                # Build per-row data context: {canonical_name: cell_value}
                row_data = {cn: read_cell_rc(ws, row, cl) for cn, cl in column_map.items() if cl}
                # Override with the value from this specific column so that
                # duplicate canonical names resolve to the correct cell.
                row_data[canonical] = read_cell_rc(ws, row, col_letter)
                try:
                    results = [evaluate_json_logic(e.get("rule", e), row_data) for e in entries]
                    result = any(results) if combine_op == "or" else all(results)
                    if not result:
                        failed_descs = [
                            e.get("description", "")
                            for e, r in zip(entries, results)
                            if not r and e.get("description")
                        ]
                        detail = "; ".join(failed_descs) if failed_descs else "règle non satisfaite"
                        issues.append(
                            make_issue(
                                "error",
                                cell_ref(col_letter, row),
                                f"Ligne {row}, colonne "
                                f"'{col_entry.get('expected', canonical)}' "
                                f"({col_letter}): {detail}",
                                mod_key,
                                "data",
                            )
                        )
                except Exception as exc:
                    issues.append(
                        make_issue(
                            "warning",
                            cell_ref(col_letter, row),
                            f"Ligne {row}: Erreur JsonLogic colonne '{canonical}': {exc}",
                            mod_key,
                            "data",
                        )
                    )

    # Validate module-level json_logic rules
    for mod_key, mod_def in schema.get("header_modules", {}).items():
        if mod_key.startswith("_") or not isinstance(mod_def, dict):
            continue
        mod_validation = mod_def.get("validation", {})
        mod_jl = mod_validation.get("json_logic")
        if not mod_jl:
            continue
        mod_entries = mod_jl if isinstance(mod_jl, list) else [mod_jl]
        mod_combine = mod_validation.get("json_logic_op", "and")
        mod_cols = module_col_letters.get(mod_key, [])
        start_col = mod_cols[0] if mod_cols else None
        for row in data_rows:
            row_data = {cn: read_cell_rc(ws, row, cl) for cn, cl in column_map.items() if cl}
            try:
                results = [evaluate_json_logic(e.get("rule", e), row_data) for e in mod_entries]
                result = any(results) if mod_combine == "or" else all(results)
                if not result:
                    failed_descs = [
                        e.get("description", "")
                        for e, r in zip(mod_entries, results)
                        if not r and e.get("description")
                    ]
                    detail = (
                        "; ".join(failed_descs)
                        if failed_descs
                        else "règle de module non satisfaite"
                    )
                    issues.append(
                        make_issue(
                            "error",
                            cell_ref(start_col or "A", row),
                            f"Ligne {row}, module '{mod_key}': {detail}",
                            mod_key,
                            "data",
                        )
                    )
            except Exception as exc:
                issues.append(
                    make_issue(
                        "warning",
                        cell_ref(start_col or "A", row),
                        f"Ligne {row}: Erreur JsonLogic module '{mod_key}': {exc}",
                        mod_key,
                        "data",
                    )
                )

    return issues


def _iter_all_columns(mod_def: dict):
    """Yield all fixed column defs (including group sub-columns) from a module."""
    for col in mod_def.get("columns", []):
        pos = col.get("position", "fixed")
        if pos == "fixed":
            yield col
        elif pos == "group":
            for sub in col.get("sub_columns", []):
                if sub.get("position") == "fixed":
                    yield sub


def find_data_rows(ws, schema: dict) -> list[int]:
    """Return list of data row indices up to the first empty row."""
    # v2 sections-based schema
    if "sections" in schema and "table_definitions" not in schema:
        start_row = _get_v2_data_start_row(schema)
    else:
        records_def = schema.get("table_definitions", {}).get("records_data", {})
        start_row = records_def.get("data_row_start", 17)
    rows = []
    for row in range(start_row, ws.max_row + 1):
        if _is_row_empty(ws, row):
            break
        rows.append(row)
    return rows


def _is_row_empty(ws, row: int) -> bool:
    """Check if a row is completely empty across columns B to BF."""
    return all(ws.cell(row=row, column=col).value is None for col in range(2, 59))


# ---------------------------------------------------------------------------
# v2 sections-based schema helpers
# ---------------------------------------------------------------------------


def _get_v2_data_start_row(schema: dict) -> int:
    """Extract data start row from the first records section in a v2 schema."""
    for section in schema.get("sections", []):
        if section.get("layout") == "records":
            return section.get("data", {}).get("start_row", 2)
    return 2


def is_v2_schema(schema: dict) -> bool:
    """Check if a schema uses the v2 sections-based format."""
    return "sections" in schema and "label_fields" not in schema


def validate_sections_headers(
    ws,
    schema: dict,
    config: dict,
) -> tuple[list[dict], dict]:
    """Validate headers for v2 sections-based schemas.

    Returns (issues, column_map) where column_map maps canonical_name -> col_letter.
    """
    issues = []
    column_map = {}

    for section in schema.get("sections", []):
        headers = section.get("headers", {})
        for col_def in headers.get("columns", []):
            col_type = col_def.get("type", "single")

            if col_type == "single":
                _validate_v2_single_header(ws, col_def, config, issues, column_map)
            elif col_type == "group":
                _validate_v2_group_header(ws, col_def, config, issues, column_map)

    return issues, column_map


def _validate_v2_single_header(
    ws,
    col_def: dict,
    config: dict,
    issues: list,
    column_map: dict,
):
    """Validate a single-column header in v2 schema."""
    col_letter = col_def.get("column")
    row = col_def.get("header_row", 1)
    expected = col_def.get("expected", "")
    canonical = col_def.get("canonical_name", "")

    if not col_letter or not expected:
        return

    column_map[canonical] = col_letter
    actual = read_cell_rc(ws, row, col_letter)

    if actual is None:
        issues.append(
            make_issue(
                "error",
                cell_ref(col_letter, row),
                f"En-tête manquant en {col_letter}{row}. Attendu : '{expected}'.",
                canonical,
                "header",
            )
        )
        return

    match = text_matches(str(actual).strip(), expected, [], config)
    if not match:
        issues.append(
            make_issue(
                "warning",
                cell_ref(col_letter, row),
                f"En-tête en {col_letter}{row} ('{actual}') ne correspond pas à '{expected}'.",
                canonical,
                "header",
            )
        )


def _validate_v2_group_header(
    ws,
    col_def: dict,
    config: dict,
    issues: list,
    column_map: dict,
):
    """Validate a group header (parent + sub_columns) in v2 schema."""
    parent = col_def.get("parent", {})
    parent_row = parent.get("header_row", 1)
    parent_expected = parent.get("expected", "")
    parent_col = parent.get("start_col")

    if parent_col and parent_expected:
        actual = read_cell_rc(ws, parent_row, parent_col)
        if actual is not None:
            match = text_matches(str(actual).strip(), parent_expected, [], config)
            if not match:
                issues.append(
                    make_issue(
                        "warning",
                        cell_ref(parent_col, parent_row),
                        f"En-tête parent en {parent_col}{parent_row} ('{actual}') "
                        f"ne correspond pas à '{parent_expected}'.",
                        parent.get("canonical_name", ""),
                        "header",
                    )
                )

    for sub in col_def.get("sub_columns", []):
        sub_col = sub.get("column")
        sub_row = sub.get("header_row", 2)
        sub_expected = sub.get("expected", "")
        sub_canonical = sub.get("canonical_name", "")

        if not sub_col or not sub_expected:
            continue

        # Prefix canonical with parent canonical to avoid collisions
        parent_canonical = parent.get("canonical_name", "")
        full_canonical = (
            f"{parent_canonical}__{sub_canonical}" if parent_canonical else sub_canonical
        )
        column_map[full_canonical] = sub_col

        actual = read_cell_rc(ws, sub_row, sub_col)
        if actual is None:
            issues.append(
                make_issue(
                    "error",
                    cell_ref(sub_col, sub_row),
                    f"En-tête manquant en {sub_col}{sub_row}. Attendu : '{sub_expected}'.",
                    full_canonical,
                    "header",
                )
            )
        else:
            match = text_matches(str(actual).strip(), sub_expected, [], config)
            if not match:
                issues.append(
                    make_issue(
                        "warning",
                        cell_ref(sub_col, sub_row),
                        f"En-tête en {sub_col}{sub_row} ('{actual}') ne correspond pas à '{sub_expected}'.",
                        full_canonical,
                        "header",
                    )
                )


def _validate_data_rows_v2(
    ws,
    schema: dict,
    column_map: dict,
) -> list[dict]:
    """Validate data rows for v2 sections-based schemas.

    Checks that required columns have non-empty values and that value types match.
    """
    issues = []

    for section in schema.get("sections", []):
        if section.get("layout") != "records":
            continue

        start_row = section.get("data", {}).get("start_row", 2)
        end_rule = section.get("data", {}).get("end_rule", "first_empty_row")
        primary_key = section.get("headers", {}).get("primary_key")

        # Find data rows
        data_rows = []
        for row in range(start_row, ws.max_row + 1):
            if end_rule == "first_empty_row" and _is_row_empty(ws, row):
                break
            data_rows.append(row)

        if not data_rows:
            issues.append(
                make_issue(
                    "warning",
                    f"A{start_row}",
                    f"Aucune ligne de données trouvée à partir de la ligne {start_row}.",
                    section.get("key", ""),
                    "data",
                )
            )
            continue

        # Validate required columns and primary key
        for col_def in section.get("headers", {}).get("columns", []):
            col_type = col_def.get("type", "single")

            if col_type == "single":
                _validate_v2_data_column(ws, col_def, data_rows, primary_key, issues)
            elif col_type == "group":
                for sub in col_def.get("sub_columns", []):
                    _validate_v2_data_column(ws, sub, data_rows, primary_key, issues)

    return issues


def _validate_v2_data_column(
    ws,
    col_def: dict,
    data_rows: list[int],
    primary_key: str | None,
    issues: list,
):
    """Validate data cells for a single column definition."""
    col_letter = col_def.get("column")
    canonical = col_def.get("canonical_name", "")
    required = col_def.get("required", False)
    value_type = col_def.get("value_type", "string")
    is_pk = canonical == primary_key

    if not col_letter:
        return

    for row in data_rows:
        val = read_cell_rc(ws, row, col_letter)
        is_empty = val is None or (isinstance(val, str) and not val.strip())

        if is_empty:
            if is_pk or required:
                issues.append(
                    make_issue(
                        "error",
                        cell_ref(col_letter, row),
                        f"Valeur manquante en {col_letter}{row} "
                        f"(colonne {'clé primaire ' if is_pk else 'requise '}"
                        f"'{col_def.get('expected', canonical)}').",
                        canonical,
                        "data",
                    )
                )
            continue

        # Type validation
        if value_type == "number" and not isinstance(val, (int, float)):
            try:
                float(val)
            except (ValueError, TypeError):
                issues.append(
                    make_issue(
                        "warning",
                        cell_ref(col_letter, row),
                        f"Valeur en {col_letter}{row} ('{val}') devrait être un nombre.",
                        canonical,
                        "data",
                    )
                )

        elif value_type == "percentage" and not isinstance(val, (int, float)):
            try:
                float(str(val).rstrip("%"))
            except (ValueError, TypeError):
                issues.append(
                    make_issue(
                        "warning",
                        cell_ref(col_letter, row),
                        f"Valeur en {col_letter}{row} ('{val}') devrait être un pourcentage.",
                        canonical,
                        "data",
                    )
                )
