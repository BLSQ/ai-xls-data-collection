"""OpenHEXA Pipeline: Excel Data Ingestion.

Reads a validated Excel file together with the extraction_guide.json
produced by the aedes-xls-validation pipeline, then writes two tables
to the workspace database:

  * program_metadata  — one row per sheet, key-value pairs from modules 1 & 2
  * program_data      — wide format: one row per Excel row (intitulé),
                         all columns from modules 3-8 as database columns
"""

import json
import uuid
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openhexa.sdk import File, current_run, parameter, pipeline, workspace
from openpyxl.utils import column_index_from_string
from sqlalchemy import create_engine, inspect, text

GUIDE_FILENAME = "extraction_guide.json"


@pipeline("xls-ingest", timeout=3600)
@parameter(
    "excel_file",
    name="Excel File Path",
    type=File,
    required=True,
    help="Path to the validated Excel file.",
)
@parameter(
    "contact_email",
    name="Contact Email",
    type=str,
    required=False,
    help="Email of the user submitting the data.",
)
def xls_ingest(excel_file: File, contact_email: str = ""):
    """Ingest a validated Excel file into the workspace database."""
    current_run.log_info(f"Starting ingestion of: {excel_file}")

    excel_path = excel_file.path
    guide_path = Path(workspace.files_path) / GUIDE_FILENAME

    # Load extraction guide
    current_run.log_info(f"Loading extraction guide from: {guide_path}")
    with guide_path.open(encoding="utf-8") as f:
        guide = json.load(f)

    # Load workbook
    current_run.log_info(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    # Duplicate rules from the extraction guide (set by the webapp)
    dup_rules = guide.get("duplicate_rules", {})
    dup_mode = dup_rules.get("mode", "no_check")
    per_user = dup_rules.get("per_user_submission", False)
    metadata_keys = [k["canonical_name"] for k in dup_rules.get("metadata_keys", [])]
    data_keys = [k["canonical_name"] for k in dup_rules.get("data_keys", [])]
    skip_delete = dup_mode == "no_check"

    current_run.log_info(
        f"Duplicate rules: mode={dup_mode}, per_user={per_user}, "
        f"metadata_keys={metadata_keys}, data_keys={data_keys}"
    )

    metadata_rows = []
    data_rows = []
    submission_datetime = datetime.now(UTC).isoformat()

    for sheet_name, sheet_guide in guide.get("sheets", {}).items():
        if sheet_name not in wb.sheetnames:
            current_run.log_warning(
                f"Sheet '{sheet_name}' in guide but not in workbook — skipping."
            )
            continue

        current_run.log_info(f"Ingesting sheet: '{sheet_name}'")
        ws = wb[sheet_name]

        # --- program_metadata ---
        meta_record = extract_metadata(ws, sheet_name, sheet_guide["metadata"])
        meta_record["contact_email"] = contact_email or ""
        meta_record["submission_datetime"] = submission_datetime

        # Compute entity_id based on duplicate rules
        entity_id = _compute_entity_id(
            dup_mode, meta_record, metadata_keys, per_user, contact_email,
        )
        meta_record["entity_id"] = entity_id
        metadata_rows.append(meta_record)

        # --- program_data ---
        sheet_data = extract_data(ws, sheet_name, sheet_guide["data"])
        for row in sheet_data:
            row["entity_id"] = entity_id
            if dup_mode == "full" and data_keys:
                row["row_id"] = _compute_row_id(entity_id, row, data_keys)
        data_rows.extend(sheet_data)

    # Build DataFrames
    df_meta = pd.DataFrame(metadata_rows)
    df_data = pd.DataFrame(data_rows)

    current_run.log_info(
        f"program_metadata: {len(df_meta)} rows, program_data: {len(df_data)} rows"
    )

    # Upsert to database (append new rows, handle schema evolution)
    engine = create_engine(workspace.database_url)

    meta_result = upsert_table(df_meta, "program_metadata", engine, skip_delete=skip_delete)
    current_run.log_info(f"program_metadata: {meta_result}")

    data_result = upsert_table(
        df_data, "program_data", engine,
        skip_delete=skip_delete,
        id_column="row_id" if (dup_mode == "full" and data_keys) else "entity_id",
    )
    current_run.log_info(f"program_data: {data_result}")

    current_run.log_info("Ingestion complete.")


# ---------------------------------------------------------------------------
# Metadata extraction (modules 1 & 2)
# ---------------------------------------------------------------------------


def extract_metadata(ws, sheet_name: str, meta_guide: dict) -> dict:
    """Extract one metadata record (modules 1 & 2) from a sheet."""
    record = {"sheet_name": sheet_name}

    # Simple fields (nom_projet, date_debut, etc.)
    for field in meta_guide.get("fields", []):
        canonical = field["canonical_name"]
        cell_ref = field["value_cell"]
        raw = ws[cell_ref].value
        record[canonical] = _coerce(raw, field.get("type", "string"))

    # Grouped sections (repeating column groups like donors, bailleurs, etc.)
    for section_name, section_info in meta_guide.get("grouped_sections", {}).items():
        col_letters = section_info.get("active_columns") or section_info.get("columns", [])
        fields_per_col = section_info.get("fields_per_column", [])

        for i, col_letter in enumerate(col_letters, start=1):
            col_idx = column_index_from_string(col_letter)
            for fdef in fields_per_col:
                # Strip the section prefix if already present, then rebuild as section_N_field
                base_name = fdef["canonical_name"]
                if base_name.startswith(f"{section_name}_"):
                    base_name = base_name[len(f"{section_name}_") :]
                canonical = f"{section_name}_{i}_{base_name}"
                raw = ws.cell(row=fdef["row"], column=col_idx).value
                record[canonical] = _coerce(raw, fdef.get("type", "string"))

    return record


# ---------------------------------------------------------------------------
# Data extraction (modules 3-8, wide format)
# ---------------------------------------------------------------------------


def extract_data(ws, sheet_name: str, data_guide: dict) -> list[dict]:
    """Extract data rows in wide format from a sheet.

    One output row per Excel row.  Every canonical column in the extraction
    guide becomes a database column under its schema-defined canonical name.

        sheet_name | row_number | budget_line_item | total_line | 2024 | ...
    """
    start = data_guide.get("start_row", 17)
    end = data_guide.get("end_row", 17)
    columns = data_guide.get("columns", {})

    # Resolve all column indices from the extraction guide
    col_specs = [
        {
            "canonical": canonical,
            "col_idx": column_index_from_string(spec["col_letter"]),
            "value_type": spec.get("value_type", "string"),
        }
        for canonical, spec in columns.items()
    ]

    rows = []
    for row_num in range(start, end + 1):
        record = {
            "sheet_name": sheet_name,
            "row_number": row_num,
        }

        for spec in col_specs:
            raw = ws.cell(row=row_num, column=spec["col_idx"]).value
            record[spec["canonical"]] = _coerce(raw, spec["value_type"])

        rows.append(record)

    return rows


# ---------------------------------------------------------------------------
# Database upsert (append with schema evolution)
# ---------------------------------------------------------------------------


def _truncate_col_names(df: pd.DataFrame, max_len: int = 63) -> pd.DataFrame:
    """Truncate column names to PostgreSQL's identifier limit (63 chars).

    If truncation causes duplicates, append a numeric suffix.
    """
    seen: dict[str, int] = {}
    new_names = []
    for col in df.columns:
        short = col[:max_len]
        if short in seen:
            seen[short] += 1
            # Make room for the suffix (e.g. "_2")
            suffix = f"_{seen[short]}"
            short = col[: max_len - len(suffix)] + suffix
        else:
            seen[short] = 1
        new_names.append(short)
    if list(df.columns) != new_names:
        df = df.copy()
        df.columns = new_names
    return df


def upsert_table(
    df: pd.DataFrame,
    table_name: str,
    engine,
    *,
    skip_delete: bool = False,
    id_column: str = "entity_id",
) -> str:
    """Append rows to a database table, handling schema evolution.

    1. If table does not exist → create it.
    2. If table exists → ALTER TABLE to add any new columns (as TEXT, NULL).
    3. DELETE existing rows whose *id_column* matches the incoming data
       (skipped when *skip_delete* is True — ``no_check`` mode).
    4. APPEND the new rows.

    Returns a human-readable summary string.
    """
    if df.empty:
        return "skipped (no rows)"

    # Truncate column names to PostgreSQL's 63-char identifier limit
    df = _truncate_col_names(df)

    inspector = inspect(engine)
    existing_tables = inspector.get_table_names()

    if table_name not in existing_tables:
        df.to_sql(table_name, engine, if_exists="fail", index=False)
        return f"created ({len(df)} rows)"

    # --- Schema evolution: add new columns ---
    existing_cols = {col["name"] for col in inspector.get_columns(table_name)}
    new_cols = [c for c in df.columns if c not in existing_cols]

    if new_cols:
        with engine.begin() as conn:
            for col in new_cols:
                sql_type = _pandas_dtype_to_sql(df[col])
                conn.execute(text(f'ALTER TABLE "{table_name}" ADD COLUMN "{col}" {sql_type}'))
        current_run.log_info(f"  {table_name}: added {len(new_cols)} new column(s): {new_cols}")

    # --- Delete existing rows with matching id_column ---
    deleted = 0
    if not skip_delete and id_column in df.columns and id_column in existing_cols:
        ids = df[id_column].dropna().unique().tolist()
        if ids:
            with engine.begin() as conn:
                result = conn.execute(
                    text(f'DELETE FROM "{table_name}" WHERE "{id_column}" = ANY(:ids)'),
                    {"ids": ids},
                )
                deleted = result.rowcount

    # --- Append new rows ---
    df.to_sql(table_name, engine, if_exists="append", index=False)

    return f"upserted ({len(df)} added, {deleted} replaced, {len(new_cols)} new columns)"


def _pandas_dtype_to_sql(series: pd.Series) -> str:
    """Map a pandas Series dtype to a PostgreSQL type for ALTER TABLE."""
    dtype = series.dtype
    if pd.api.types.is_float_dtype(dtype):
        return "DOUBLE PRECISION"
    if pd.api.types.is_integer_dtype(dtype):
        return "BIGINT"
    if pd.api.types.is_bool_dtype(dtype):
        return "BOOLEAN"
    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "TIMESTAMP"
    return "TEXT"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _compute_entity_id(
    mode: str,
    meta_record: dict,
    metadata_keys: list[str],
    per_user: bool,
    contact_email: str,
) -> str:
    """Compute the entity_id for a metadata record.

    - ``no_check``: random UUID v4 (every submission is unique).
    - ``metadata_unique`` / ``full``: deterministic UUID v5 built from
      the values of the selected *metadata_keys*.  When *per_user* is
      True, *contact_email* is prepended so each user gets their own
      submission identity.
    """
    if mode == "no_check" or not metadata_keys:
        if mode != "no_check":
            current_run.log_warning(
                f"Duplicate mode '{mode}' but metadata_keys is empty — "
                "falling back to random UUID (no deduplication)."
            )
        return str(uuid.uuid4())

    parts: list[str] = []
    if per_user:
        parts.append(contact_email or "")
    for key in sorted(metadata_keys):
        val = meta_record.get(key)
        parts.append("" if val is None else str(val))
    seed = "::".join(parts)
    return str(uuid.uuid5(uuid.NAMESPACE_URL, seed))


def _compute_row_id(entity_id: str, row: dict, data_keys: list[str]) -> str:
    """Compute a per-row unique ID for ``full`` duplicate mode.

    Combines the *entity_id* with the values of the selected *data_keys*
    columns to produce a deterministic UUID v5 for each data row.
    """
    parts = [entity_id]
    for key in sorted(data_keys):
        val = row.get(key)
        parts.append("" if val is None else str(val))
    return str(uuid.uuid5(uuid.NAMESPACE_URL, "::".join(parts)))


def _coerce(raw, value_type: str):
    """Coerce a raw cell value to a string representation for storage."""
    if raw is None:
        return None
    # Dates: openpyxl returns datetime objects
    if hasattr(raw, "isoformat"):
        return raw.isoformat()
    if value_type == "number":
        try:
            return float(raw)
        except (ValueError, TypeError):
            return str(raw).strip() if raw is not None else None
    return str(raw).strip() if raw is not None else None


if __name__ == "__main__":
    xls_ingest()
