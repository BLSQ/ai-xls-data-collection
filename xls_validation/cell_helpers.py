"""Low-level cell reading and issue-building utilities.

Thin wrappers around openpyxl cell access and a standard issue dict
factory used by all section validators.

Pure functions — no side effects, no OpenHEXA dependency.
"""

from __future__ import annotations

from openpyxl.utils import column_index_from_string


def read_cell(worksheet, cell_reference: str):
    """Read a cell value from a worksheet by its A1 reference.

    Args:
        worksheet: An openpyxl ``Worksheet``.
        cell_reference: An A1-notation reference such as ``"B3"``.

    Returns:
        The raw cell value (string, number, datetime, or None).
    """
    return worksheet[cell_reference].value


def read_cell_by_row_column(worksheet, row: int, column_letter: str):
    """Read a cell value by row number and column letter.

    Args:
        worksheet: An openpyxl ``Worksheet``.
        row: The 1-based row number.
        column_letter: The column letter(s) such as ``"C"`` or ``"AB"``.

    Returns:
        The raw cell value.
    """
    column_index = column_index_from_string(column_letter)
    return worksheet.cell(row=row, column=column_index).value


def cell_reference(column_letter: str, row: int) -> str:
    """Build an A1-notation cell reference from column letter and row number.

    Args:
        column_letter: The column letter(s).
        row: The 1-based row number.

    Returns:
        A string such as ``"C7"``.
    """
    return f"{column_letter}{row}"


def is_row_empty(worksheet, row: int, max_column: int = 58) -> bool:
    """Check whether a row is completely empty across columns B to the given max.

    Args:
        worksheet: An openpyxl ``Worksheet``.
        row: The 1-based row number.
        max_column: The last 1-based column index to check (default 58 = BF).

    Returns:
        True if every cell in the range is None.
    """
    return all(
        worksheet.cell(row=row, column=col).value is None
        for col in range(2, max_column + 1)
    )


def find_data_rows(worksheet, start_row: int) -> list[int]:
    """Return a list of data row indices from *start_row* to the first empty row.

    Args:
        worksheet: An openpyxl ``Worksheet``.
        start_row: The first row to check.

    Returns:
        A list of 1-based row numbers that contain data.
    """
    rows: list[int] = []
    for row in range(start_row, worksheet.max_row + 1):
        if is_row_empty(worksheet, row):
            break
        rows.append(row)
    return rows


def make_issue(
    severity: str,
    cell: str,
    message: str,
    field_reference: str = "",
    group: str = "",
) -> dict:
    """Build a standardized validation issue dictionary.

    Args:
        severity: One of ``"error"``, ``"warning"``, ``"info"``.
        cell: The A1-notation cell reference where the issue was found.
        message: A human-readable description of the issue.
        field_reference: The schema field key this issue relates to.
        group: The validation group (``"label"``, ``"value"``, ``"header"``,
            ``"data"``).

    Returns:
        A dict with keys ``severity``, ``cell``, ``message``,
        ``field_ref``, ``group``.
    """
    return {
        "severity": severity,
        "cell": cell,
        "message": message,
        "field_ref": field_reference,
        "group": group,
    }
