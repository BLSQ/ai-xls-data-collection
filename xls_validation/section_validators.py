"""Polymorphic section validators for Excel schema validation.

Provides an abstract base class :class:`SectionValidator` with two abstract
methods — ``validate_structure()`` and ``validate_data()`` — and three
concrete implementations:

- :class:`KeyValueValidator`          — single-cell label + value pairs
- :class:`GroupedKeyValueValidator`   — columns × rows grouped entries
- :class:`RecordsValidator`           — tabular headers + data rows

All validators are schema-format agnostic: they accept pre-extracted schema
slices so the orchestrator (``pipeline.py``) handles v1/v2 dispatch.

Pure validation logic — no OpenHEXA SDK dependency.
"""

from __future__ import annotations

import re
from abc import ABC, abstractmethod
from datetime import datetime

from openpyxl.utils import column_index_from_string, get_column_letter

from cell_helpers import (
    cell_reference,
    find_data_rows,
    make_issue,
    read_cell,
    read_cell_by_row_column,
)
from json_logic import evaluate_json_logic
from text_helpers import text_matches, to_canonical


# ---------------------------------------------------------------------------
# Abstract base
# ---------------------------------------------------------------------------


class SectionValidator(ABC):
    """Abstract base class for section validators.

    Subclasses must implement ``validate_structure()`` (labels, headers) and
    ``validate_data()`` (values, data rows).
    """

    def __init__(self, worksheet, config: dict) -> None:
        """Initialise the validator.

        Args:
            worksheet: An openpyxl ``Worksheet`` to validate.
            config: The ``validation_config.label_comparison`` dict from the
                schema, controlling normalization behavior.
        """
        self.worksheet = worksheet
        self.config = config

    @abstractmethod
    def validate_structure(self) -> list[dict]:
        """Validate structural elements (labels exist, headers match).

        Returns:
            A list of issue dicts.
        """

    @abstractmethod
    def validate_data(self) -> list[dict]:
        """Validate data values (types, required fields, enums, rules).

        Returns:
            A list of issue dicts.
        """

    def process(self) -> list[dict]:
        """Run structure validation then data validation.

        Convenience method that calls both abstract methods in order.

        Returns:
            A combined list of issue dicts.
        """
        issues = self.validate_structure()
        issues.extend(self.validate_data())
        return issues


# ---------------------------------------------------------------------------
# KeyValueValidator — label + single-cell value validation
# ---------------------------------------------------------------------------


class KeyValueValidator(SectionValidator):
    """Validates key-value label/value pairs.

    Handles ``label_fields`` (labels at expected cells) and ``value_fields``
    (single-cell value entries with type/enum/regex/json_logic checks).
    """

    def __init__(
        self,
        worksheet,
        config: dict,
        label_fields: dict,
        value_fields: dict,
    ) -> None:
        """Initialise the key-value validator.

        Args:
            worksheet: An openpyxl ``Worksheet``.
            config: Normalization config dict.
            label_fields: The ``schema["label_fields"]`` dict (may be empty).
            value_fields: The ``schema["value_fields"]`` dict (may be empty).
        """
        super().__init__(worksheet, config)
        self.label_fields = label_fields or {}
        self.value_fields = value_fields or {}

    # -- Structure: labels --------------------------------------------------

    def validate_structure(self) -> list[dict]:
        """Validate that all label_fields are present at their expected cells.

        Returns:
            A list of issue dicts for missing or mismatched labels.
        """
        issues: list[dict] = []

        for module_key, module_definition in self.label_fields.items():
            if module_key.startswith("_"):
                continue

            # Entries with "items" list (e.g. *_columns entries)
            if "items" in module_definition:
                for item in module_definition["items"]:
                    self._check_label(item, issues, module_key)
                continue

            for field_key, field_definition in module_definition.items():
                if field_key.startswith("_"):
                    continue
                if not isinstance(field_definition, dict) or "cell" not in field_definition:
                    continue
                self._check_label(
                    field_definition, issues, f"{module_key}.{field_key}"
                )

        return issues

    def _check_label(
        self,
        field_definition: dict,
        issues: list[dict],
        field_reference: str,
    ) -> None:
        """Check a single label field against the worksheet.

        If the expected cell is empty, scans ±3 columns on the same row
        for a shifted label.

        Args:
            field_definition: A dict with ``cell``, ``expected``, and optional
                ``known_variants``.
            issues: The issues list to append to.
            field_reference: The schema key path for issue attribution.
        """
        cell = field_definition["cell"]
        if not cell:
            return
        expected = field_definition["expected"]
        variants = field_definition.get("known_variants", [])
        actual = read_cell(self.worksheet, cell)

        if actual is not None:
            match = text_matches(str(actual), expected, variants, self.config)
            if match:
                return
            issues.append(
                make_issue(
                    "warning",
                    cell,
                    f"Le label en {cell} ('{actual}') ne correspond pas "
                    f"exactement à '{expected}'. Vérifiez la cellule.",
                    field_reference,
                    "label",
                )
            )
            return

        # Cell is empty — scan ±3 columns on the same row
        match_result = re.match(r"^([A-Z]+)(\d+)$", cell)
        if match_result:
            column_index = column_index_from_string(match_result.group(1))
            row = int(match_result.group(2))
            for offset in (-1, 1, -2, 2, -3, 3):
                try_index = column_index + offset
                if try_index < 1:
                    continue
                try_column = get_column_letter(try_index)
                try_value = read_cell_by_row_column(self.worksheet, row, try_column)
                if try_value is None:
                    continue
                match = text_matches(str(try_value), expected, variants, self.config)
                if match:
                    issues.append(
                        make_issue(
                            "info",
                            f"{try_column}{row}",
                            f"Label '{expected}' trouvé en {try_column}{row} "
                            f"au lieu de {cell} (décalage de {offset} colonne(s)).",
                            field_reference,
                            "label",
                        )
                    )
                    return

        issues.append(
            make_issue(
                "error",
                cell,
                f"Label attendu manquant en {cell}. Attendu : '{expected}'.",
                field_reference,
                "label",
            )
        )

    # -- Data: single-cell values -------------------------------------------

    def validate_data(self) -> list[dict]:
        """Validate single-cell value entries (type, enum, regex, json_logic).

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []

        for section_key, section in self.value_fields.items():
            if section_key.startswith("_") or not isinstance(section, dict):
                continue
            for field_key, field_definition in section.items():
                if field_key.startswith("_") or not isinstance(field_definition, dict):
                    continue
                # Skip grouped entries (handled by GroupedKeyValueValidator)
                if "columns" in field_definition and "rows" in field_definition:
                    continue
                if "cell" in field_definition:
                    issues.extend(
                        self._validate_single_value(field_key, field_definition)
                    )

        return issues

    def _validate_single_value(
        self,
        field_key: str,
        field_definition: dict,
    ) -> list[dict]:
        """Validate a single-cell value entry.

        Checks null/required, type (number/date/string), enum, regex,
        and cell-level json_logic rules.

        Args:
            field_key: The field's key in the schema.
            field_definition: The field definition dict.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        cell = field_definition["cell"]
        value = read_cell(self.worksheet, cell)
        label_display = field_definition.get("label_display", field_key)
        field_reference = field_definition.get("label_ref", field_key)
        required = field_definition.get("required", False)
        field_type = field_definition.get("type", "string")
        validation = field_definition.get("validation", {})
        message = field_definition.get("message", "")

        # -- Null check --
        if value is None or (isinstance(value, str) and not value.strip()):
            severity = "error" if required else "info"
            issues.append(make_issue(severity, cell, message, field_reference, "value"))
            return issues

        # -- Type checks --
        example = field_definition.get("example", "")
        example_hint = f" Exemple attendu : '{example}'." if example else ""

        if field_type == "number":
            issues.extend(
                self._check_number_value(
                    cell, value, label_display, field_reference, example_hint,
                    message, validation,
                )
            )
        elif field_type == "date":
            if not isinstance(value, datetime) and not _is_date_string(value):
                msg = (
                    f"Le champ '{label_display}' ({cell}) doit contenir une "
                    f"date valide. Valeur actuelle : '{value}'.{example_hint}"
                ) if example_hint else message
                issues.append(make_issue("error", cell, msg, field_reference, "value"))
        elif field_type == "string":
            issues.extend(
                self._check_string_value(
                    cell, value, label_display, field_reference, example_hint,
                    message, validation,
                )
            )

        # -- Cell-level json_logic rules --
        issues.extend(
            self._check_json_logic(cell, value, label_display, field_reference, validation)
        )

        return issues

    @staticmethod
    def _check_number_value(
        cell: str,
        value,
        label_display: str,
        field_reference: str,
        example_hint: str,
        message: str,
        validation: dict,
    ) -> list[dict]:
        """Validate a number-type cell value."""
        issues: list[dict] = []
        try:
            numeric_value = float(value)
            if "min" in validation and numeric_value < validation["min"]:
                issues.append(make_issue("error", cell, message, field_reference, "value"))
        except (ValueError, TypeError):
            issues.append(
                make_issue(
                    "error",
                    cell,
                    f"Le champ '{label_display}' ({cell}) doit être "
                    f"un nombre. Valeur actuelle : '{value}'.{example_hint}",
                    field_reference,
                    "value",
                )
            )
        return issues

    @staticmethod
    def _check_string_value(
        cell: str,
        value,
        label_display: str,
        field_reference: str,
        example_hint: str,
        message: str,
        validation: dict,
    ) -> list[dict]:
        """Validate a string-type cell value (min_length, regex, enum)."""
        issues: list[dict] = []
        string_value = str(value).strip()
        if "min_length" in validation and len(string_value) < validation["min_length"]:
            issues.append(make_issue("error", cell, message, field_reference, "value"))
        if "regex" in validation and not re.match(validation["regex"], string_value):
            msg = (
                f"Le champ '{label_display}' ({cell}) ne correspond pas "
                f"au format attendu. Valeur actuelle : '{string_value}'.{example_hint}"
            ) if example_hint else message
            issues.append(make_issue("error", cell, msg, field_reference, "value"))
        if "enum" in validation and string_value not in validation["enum"]:
            issues.append(
                make_issue(
                    "error",
                    cell,
                    f"Le champ '{label_display}' ({cell}) contient "
                    f"'{string_value}'. Valeurs acceptées : "
                    f"{validation['enum']}.{example_hint}",
                    field_reference,
                    "value",
                )
            )
        return issues

    @staticmethod
    def _check_json_logic(
        cell: str,
        value,
        label_display: str,
        field_reference: str,
        validation: dict,
    ) -> list[dict]:
        """Evaluate cell-level json_logic rules."""
        issues: list[dict] = []
        json_logic_definition = validation.get("json_logic")
        if not json_logic_definition:
            return issues

        entries = (
            json_logic_definition
            if isinstance(json_logic_definition, list)
            else [json_logic_definition]
        )
        combine_operator = validation.get("json_logic_op", "and")
        cell_data = {"value": value}

        try:
            results = [
                evaluate_json_logic(entry.get("rule", entry), cell_data)
                for entry in entries
            ]
            combined = any(results) if combine_operator == "or" else all(results)
            if not combined:
                failed_descriptions = [
                    entry.get("description", "")
                    for entry, result in zip(entries, results)
                    if not result and entry.get("description")
                ]
                detail = (
                    "; ".join(failed_descriptions)
                    if failed_descriptions
                    else "règle non satisfaite"
                )
                issues.append(
                    make_issue(
                        "error", cell,
                        f"Le champ '{label_display}' ({cell}): {detail}",
                        field_reference, "value",
                    )
                )
        except Exception as exc:
            issues.append(
                make_issue(
                    "warning", cell,
                    f"Erreur JsonLogic pour '{label_display}' ({cell}): {exc}",
                    field_reference, "value",
                )
            )

        return issues


# ---------------------------------------------------------------------------
# GroupedKeyValueValidator — columns × rows grouped entries
# ---------------------------------------------------------------------------


class GroupedKeyValueValidator(SectionValidator):
    """Validates grouped key-value entries (columns × rows structure).

    Each grouped entry defines a set of columns and rows with an anchor row
    that determines which columns are "active". Other rows with
    ``required_rule: if_active`` are validated only for active columns.
    """

    def __init__(
        self,
        worksheet,
        config: dict,
        grouped_definitions: list[dict],
    ) -> None:
        """Initialise the grouped key-value validator.

        Args:
            worksheet: An openpyxl ``Worksheet``.
            config: Normalization config dict.
            grouped_definitions: A list of grouped field definition dicts,
                each containing ``columns`` and ``rows`` keys.
        """
        super().__init__(worksheet, config)
        self.grouped_definitions = grouped_definitions or []

    def validate_structure(self) -> list[dict]:
        """Grouped entries have no structural validation.

        Returns:
            An empty list.
        """
        return []

    def validate_data(self) -> list[dict]:
        """Validate all grouped entries.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        for grouped_definition in self.grouped_definitions:
            issues.extend(self._validate_grouped_entry(grouped_definition))
        return issues

    def _validate_grouped_entry(self, grouped_definition: dict) -> list[dict]:
        """Validate a single grouped entry (columns × rows).

        Finds the anchor row to determine active columns, then validates
        ``if_active`` rows for those columns.

        Args:
            grouped_definition: A dict with ``columns`` and ``rows`` keys.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        columns = grouped_definition["columns"]
        rows_definition = grouped_definition.get("rows", {})

        # Find the anchor row
        anchor_row_key = None
        for row_key, row_def in rows_definition.items():
            if row_def.get("required_rule") == "anchor":
                anchor_row_key = row_key
                break

        if anchor_row_key is None:
            return issues

        anchor_row_definition = rows_definition[anchor_row_key]
        anchor_row = anchor_row_definition["row"]

        # Determine active columns (non-empty in anchor row)
        active_columns: list[str] = []
        for column in columns:
            value = read_cell_by_row_column(self.worksheet, anchor_row, column)
            if value is not None and str(value).strip():
                active_columns.append(column)

        # Anchor check: at least one column must have a value
        if not active_columns:
            issues.append(
                make_issue(
                    "error",
                    cell_reference(columns[0], anchor_row),
                    anchor_row_definition.get("message", ""),
                    anchor_row_definition.get("label_ref", ""),
                    "value",
                )
            )

        # Validate other rows for active columns
        for row_key, row_def in rows_definition.items():
            if row_key == anchor_row_key:
                continue
            if row_def.get("required_rule") != "if_active":
                continue

            row_number = row_def["row"]
            for column in active_columns:
                issues.extend(
                    self._validate_active_cell(column, row_number, row_def)
                )

        return issues

    def _validate_active_cell(
        self,
        column: str,
        row_number: int,
        row_definition: dict,
    ) -> list[dict]:
        """Validate a single cell in an active column.

        Args:
            column: The column letter.
            row_number: The row number.
            row_definition: The row definition dict from the schema.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        value = read_cell_by_row_column(self.worksheet, row_number, column)

        if value is None or (isinstance(value, str) and not value.strip()):
            issues.append(
                make_issue(
                    "error",
                    cell_reference(column, row_number),
                    f"{row_definition.get('message', '')} (colonne {column})",
                    row_definition.get("label_ref", ""),
                    "value",
                )
            )
        elif row_definition.get("type") == "number":
            try:
                numeric_value = float(value)
                min_value = row_definition.get("validation", {}).get("min")
                if min_value is not None and numeric_value < min_value:
                    issues.append(
                        make_issue(
                            "error",
                            cell_reference(column, row_number),
                            f"{row_definition.get('message', '')} (colonne {column})"
                            " — valeur négative.",
                            row_definition.get("label_ref", ""),
                            "value",
                        )
                    )
            except (ValueError, TypeError):
                example = row_definition.get("example", "")
                hint = f" Exemple attendu : '{example}'." if example else ""
                issues.append(
                    make_issue(
                        "error",
                        cell_reference(column, row_number),
                        f"{row_definition.get('label_display', '')} "
                        f"({column}{row_number}) doit être un nombre.{hint}",
                        row_definition.get("label_ref", ""),
                        "value",
                    )
                )

        return issues


# ---------------------------------------------------------------------------
# RecordsValidator — header + data row validation
# ---------------------------------------------------------------------------


class RecordsValidator(SectionValidator):
    """Validates records tables: header row(s) and data rows.

    Handles both flat-key schemas (``header_modules`` + ``table_definitions``)
    and sections-based schemas (``sections[layout=records]``). The public
    interface is identical regardless of schema format.

    After ``validate_structure()`` is called, ``column_map`` and
    ``module_columns`` are populated and available for downstream use
    (e.g. building the extraction guide).
    """

    def __init__(
        self,
        worksheet,
        config: dict,
        header_modules: dict | None = None,
        table_definitions: dict | None = None,
        sections: list[dict] | None = None,
    ) -> None:
        """Initialise the records validator.

        Provide *either* ``header_modules`` + ``table_definitions`` (flat-key
        schema) *or* ``sections`` (sections-based schema).

        Args:
            worksheet: An openpyxl ``Worksheet``.
            config: Normalization config dict.
            header_modules: The ``schema["header_modules"]`` dict (flat-key).
            table_definitions: The ``schema["table_definitions"]`` dict
                (flat-key).
            sections: A list of section dicts with ``layout: records``
                (sections-based).
        """
        super().__init__(worksheet, config)
        self.header_modules = header_modules or {}
        self.table_definitions = table_definitions or {}
        self.sections = sections or []
        self._column_map: dict[str, str] = {}
        self._module_columns: dict[str, list[str]] = {}

    @property
    def column_map(self) -> dict[str, str]:
        """Return the canonical_name → column_letter mapping.

        Populated after ``validate_structure()`` is called.
        """
        return self._column_map

    @property
    def module_columns(self) -> dict[str, list[str]]:
        """Return the module_key → list of column letters mapping.

        Populated after ``validate_structure()`` for flat-key schemas.
        """
        return self._module_columns

    # -- Structure validation -----------------------------------------------

    def validate_structure(self) -> list[dict]:
        """Validate header rows and build column mappings.

        Returns:
            A list of issue dicts for missing or mismatched headers.
        """
        if self.sections:
            return self._validate_sections_headers()
        return self._validate_module_headers()

    def _validate_module_headers(self) -> list[dict]:
        """Validate flat-key header_modules (fixed/group/dynamic columns).

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        mismatch_template = self.header_modules.get("mismatch_message_template", "")

        for module_key, module_definition in self.header_modules.items():
            if module_key.startswith("_"):
                continue
            if not isinstance(module_definition, dict) or "columns" not in module_definition:
                continue

            module_letters: list[str] = []

            # Validate module title
            title_definition = module_definition.get("title")
            if title_definition:
                issues.extend(
                    self._check_module_title(title_definition, module_key)
                )

            column_cursor = column_index_from_string(
                module_definition.get("start_column", "B")
            )

            for column_entry in module_definition["columns"]:
                position = column_entry.get("position", "fixed")

                if position == "fixed":
                    self._validate_fixed_header(
                        column_entry, mismatch_template, issues
                    )
                    column_letter = column_entry.get("column")
                    if column_letter:
                        module_letters.append(column_letter)
                        column_cursor = column_index_from_string(column_letter) + 1

                elif position == "group":
                    group_letters, column_cursor = self._validate_group_header(
                        column_entry, column_cursor, mismatch_template, issues,
                        module_key,
                    )
                    module_letters.extend(group_letters)

                elif position == "dynamic":
                    dynamic_canonicals, column_cursor = self._validate_dynamic_headers(
                        column_entry, column_cursor, issues
                    )
                    for canonical in dynamic_canonicals:
                        column_letter = self._column_map.get(canonical)
                        if column_letter:
                            module_letters.append(column_letter)

            self._module_columns[module_key] = module_letters

        return issues

    def _check_module_title(
        self,
        title_definition: dict,
        module_key: str,
    ) -> list[dict]:
        """Check a module's title cell against its expected value.

        Args:
            title_definition: A dict with ``cell`` and ``expected`` keys.
            module_key: The module key for issue attribution.

        Returns:
            A list of issue dicts (0 or 1).
        """
        issues: list[dict] = []
        title_cell = title_definition["cell"]
        title_value = read_cell(self.worksheet, title_cell)
        if title_value is not None:
            match = text_matches(
                str(title_value), title_definition["expected"], [], self.config
            )
            if not match:
                issues.append(
                    make_issue(
                        "warning",
                        title_cell,
                        f"Titre du {module_key} en {title_cell} ('{title_value}') "
                        f"ne correspond pas à '{title_definition['expected']}'.",
                        module_key,
                        "header",
                    )
                )
        return issues

    def _validate_fixed_header(
        self,
        column_definition: dict,
        mismatch_template: str,
        issues: list[dict],
    ) -> None:
        """Validate a single fixed-position header column.

        Populates ``self._column_map`` with the canonical → column mapping.

        Args:
            column_definition: The column entry dict from the schema.
            mismatch_template: A format string for mismatch messages.
            issues: The issues list to append to.
        """
        column_letter = column_definition["column"]
        row = column_definition["row"]
        expected = column_definition["expected"]
        canonical = column_definition["canonical_name"]
        variants = column_definition.get("known_variants", [])

        actual = read_cell_by_row_column(self.worksheet, row, column_letter)
        self._column_map[canonical] = column_letter

        if actual is None:
            issues.append(
                make_issue(
                    "error",
                    cell_reference(column_letter, row),
                    f"En-tête manquant en {column_letter}{row}. "
                    f"Attendu : '{expected}' (→ {canonical}).",
                    canonical,
                    "header",
                )
            )
            return

        actual_string = str(actual).strip()
        match = text_matches(actual_string, expected, variants, self.config)

        if match == "variant":
            issues.append(
                make_issue(
                    "info",
                    cell_reference(column_letter, row),
                    f"En-tête en {column_letter}{row} ('{actual_string}') "
                    f"reconnu comme variante de '{expected}'.",
                    canonical,
                    "header",
                )
            )
        elif not match:
            msg = (
                mismatch_template.format(
                    column=column_letter,
                    actual_value=actual_string,
                    expected=expected,
                    canonical_name=canonical,
                )
                if mismatch_template
                else (
                    f"En-tête en {column_letter}{row} ('{actual_string}') "
                    f"ne correspond pas à '{expected}'."
                )
            )
            issues.append(
                make_issue("warning", cell_reference(column_letter, row), msg, canonical, "header")
            )

    def _validate_group_header(
        self,
        group_definition: dict,
        column_cursor: int,
        mismatch_template: str,
        issues: list[dict],
        module_key: str,
    ) -> tuple[list[str], int]:
        """Validate a group header (parent + sub_columns).

        Args:
            group_definition: The group column entry dict.
            column_cursor: The current column cursor position.
            mismatch_template: A format string for mismatch messages.
            issues: The issues list to append to.
            module_key: The parent module key for issue attribution.

        Returns:
            A tuple of (column_letters, new_cursor).
        """
        parent = group_definition["parent"]
        parent_row = parent["row"]
        parent_expected = parent["expected"]
        parent_variants = parent.get("known_variants", [])
        column_letters: list[str] = []

        parent_column = get_column_letter(column_cursor)
        parent_value = read_cell_by_row_column(self.worksheet, parent_row, parent_column)

        if parent_value is not None:
            match = text_matches(
                str(parent_value), parent_expected, parent_variants, self.config
            )
            if match == "variant":
                issues.append(
                    make_issue(
                        "info",
                        cell_reference(parent_column, parent_row),
                        f"En-tête parent en {parent_column}{parent_row} "
                        f"('{parent_value}') reconnu comme variante de "
                        f"'{parent_expected}'.",
                        module_key,
                        "header",
                    )
                )
            elif not match:
                issues.append(
                    make_issue(
                        "warning",
                        cell_reference(parent_column, parent_row),
                        f"En-tête parent en {parent_column}{parent_row} "
                        f"('{parent_value}') ne correspond pas à "
                        f"'{parent_expected}'.",
                        module_key,
                        "header",
                    )
                )

        cursor = column_cursor
        for sub_column in group_definition.get("sub_columns", []):
            sub_position = sub_column.get("position", "fixed")

            if sub_position == "fixed":
                sub_column_letter = sub_column.get("column") or get_column_letter(cursor)
                sub_with_column = {**sub_column, "column": sub_column_letter}
                self._validate_fixed_header(
                    sub_with_column, mismatch_template, issues
                )
                column_letters.append(sub_column_letter)
                cursor = column_index_from_string(sub_column_letter) + 1

            elif sub_position == "dynamic":
                dynamic_canonicals, cursor = self._validate_dynamic_headers(
                    sub_column, cursor, issues
                )
                for canonical in dynamic_canonicals:
                    column_letter = self._column_map.get(canonical)
                    if column_letter:
                        column_letters.append(column_letter)

        return column_letters, cursor

    def _validate_dynamic_headers(
        self,
        dynamic_definition: dict,
        column_cursor: int,
        issues: list[dict],
    ) -> tuple[list[str], int]:
        """Scan consecutive cells matching a regex pattern.

        Args:
            dynamic_definition: The dynamic column entry dict.
            column_cursor: The starting column cursor.
            issues: The issues list to append to.

        Returns:
            A tuple of (canonical_names, new_cursor).
        """
        row = dynamic_definition["row"]
        pattern = dynamic_definition.get("pattern", "")
        prefix = dynamic_definition.get("canonical_prefix", "dynamic")
        min_count = dynamic_definition.get("min_count", 0)
        canonicals: list[str] = []
        cursor = column_cursor

        while cursor <= 16384:
            column_letter = get_column_letter(cursor)
            value = read_cell_by_row_column(self.worksheet, row, column_letter)

            if value is None:
                break
            value_string = str(value).strip()
            if not re.match(pattern, value_string):
                break

            normalized = to_canonical(value_string)
            canonical = f"{prefix}_{normalized}"
            self._column_map[canonical] = column_letter
            canonicals.append(canonical)
            cursor += 1

        if len(canonicals) < min_count:
            column_letter = get_column_letter(column_cursor)
            issues.append(
                make_issue(
                    "error",
                    cell_reference(column_letter, row),
                    f"Au moins {min_count} colonne(s) dynamique(s) attendue(s) "
                    f"(pattern: {pattern}) à partir de {column_letter}{row}, "
                    f"mais {len(canonicals)} trouvée(s).",
                    prefix,
                    "header",
                )
            )

        return canonicals, cursor

    # -- Sections-based header validation -----------------------------------

    def _validate_sections_headers(self) -> list[dict]:
        """Validate headers for sections-based schemas.

        Handles ``type: single`` and ``type: group`` column definitions.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []

        for section in self.sections:
            headers = section.get("headers", {})
            for column_definition in headers.get("columns", []):
                column_type = column_definition.get("type", "single")
                if column_type == "single":
                    self._validate_single_column_header(
                        column_definition, issues
                    )
                elif column_type == "group":
                    self._validate_group_column_header(
                        column_definition, issues
                    )

        return issues

    def _validate_single_column_header(
        self,
        column_definition: dict,
        issues: list[dict],
    ) -> None:
        """Validate a single-column header in a sections-based schema.

        Args:
            column_definition: The column definition dict.
            issues: The issues list to append to.
        """
        column_letter = column_definition.get("column")
        row = column_definition.get("header_row", 1)
        expected = column_definition.get("expected", "")
        canonical = column_definition.get("canonical_name", "")

        if not column_letter or not expected:
            return

        self._column_map[canonical] = column_letter
        actual = read_cell_by_row_column(self.worksheet, row, column_letter)

        if actual is None:
            issues.append(
                make_issue(
                    "error",
                    cell_reference(column_letter, row),
                    f"En-tête manquant en {column_letter}{row}. "
                    f"Attendu : '{expected}'.",
                    canonical,
                    "header",
                )
            )
            return

        match = text_matches(str(actual).strip(), expected, [], self.config)
        if not match:
            issues.append(
                make_issue(
                    "warning",
                    cell_reference(column_letter, row),
                    f"En-tête en {column_letter}{row} ('{actual}') "
                    f"ne correspond pas à '{expected}'.",
                    canonical,
                    "header",
                )
            )

    def _validate_group_column_header(
        self,
        column_definition: dict,
        issues: list[dict],
    ) -> None:
        """Validate a group header (parent + sub_columns) in sections-based schema.

        Args:
            column_definition: The group column definition dict.
            issues: The issues list to append to.
        """
        parent = column_definition.get("parent", {})
        parent_row = parent.get("header_row", 1)
        parent_expected = parent.get("expected", "")
        parent_column = parent.get("start_col")

        if parent_column and parent_expected:
            actual = read_cell_by_row_column(self.worksheet, parent_row, parent_column)
            if actual is not None:
                match = text_matches(
                    str(actual).strip(), parent_expected, [], self.config
                )
                if not match:
                    issues.append(
                        make_issue(
                            "warning",
                            cell_reference(parent_column, parent_row),
                            f"En-tête parent en {parent_column}{parent_row} "
                            f"('{actual}') ne correspond pas à '{parent_expected}'.",
                            parent.get("canonical_name", ""),
                            "header",
                        )
                    )

        parent_canonical = parent.get("canonical_name", "")
        for sub_column in column_definition.get("sub_columns", []):
            sub_column_letter = sub_column.get("column")
            sub_row = sub_column.get("header_row", 2)
            sub_expected = sub_column.get("expected", "")
            sub_canonical = sub_column.get("canonical_name", "")

            if not sub_column_letter or not sub_expected:
                continue

            full_canonical = (
                f"{parent_canonical}__{sub_canonical}"
                if parent_canonical
                else sub_canonical
            )
            self._column_map[full_canonical] = sub_column_letter

            actual = read_cell_by_row_column(self.worksheet, sub_row, sub_column_letter)
            if actual is None:
                issues.append(
                    make_issue(
                        "error",
                        cell_reference(sub_column_letter, sub_row),
                        f"En-tête manquant en {sub_column_letter}{sub_row}. "
                        f"Attendu : '{sub_expected}'.",
                        full_canonical,
                        "header",
                    )
                )
            else:
                match = text_matches(
                    str(actual).strip(), sub_expected, [], self.config
                )
                if not match:
                    issues.append(
                        make_issue(
                            "warning",
                            cell_reference(sub_column_letter, sub_row),
                            f"En-tête en {sub_column_letter}{sub_row} "
                            f"('{actual}') ne correspond pas à '{sub_expected}'.",
                            full_canonical,
                            "header",
                        )
                    )

    # -- Data row validation ------------------------------------------------

    def validate_data(self) -> list[dict]:
        """Validate data rows (primary key, enums, json_logic, types).

        Returns:
            A list of issue dicts.
        """
        if self.sections:
            return self._validate_sections_data()
        return self._validate_module_data()

    def _validate_module_data(self) -> list[dict]:
        """Validate data rows for flat-key schemas.

        Checks primary key presence, enum constraints, column-level and
        module-level json_logic rules.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        records_definition = self.table_definitions.get("records_data")
        if not records_definition:
            return issues

        row_validation = records_definition.get("row_validation", {})
        start_row = records_definition.get("data_row_start", 17)
        data_rows = find_data_rows(self.worksheet, start_row)

        if not data_rows:
            issues.append(
                make_issue(
                    "warning",
                    f"B{start_row}",
                    f"Aucune ligne de données trouvée à partir de la ligne {start_row}.",
                    "records_data",
                    "data",
                )
            )
            return issues

        # Primary key validation
        issues.extend(
            self._validate_primary_keys(row_validation, data_rows, start_row)
        )

        # Enum constraints
        for module_key, module_definition in self.header_modules.items():
            if module_key.startswith("_") or not isinstance(module_definition, dict):
                continue
            for column_entry in module_definition.get("columns", []):
                self._check_column_enum(
                    column_entry, data_rows, issues, module_key
                )

        # Column-level json_logic rules
        issues.extend(self._validate_column_json_logic(data_rows))

        # Module-level json_logic rules
        issues.extend(self._validate_module_json_logic(data_rows))

        return issues

    def _validate_primary_keys(
        self,
        row_validation: dict,
        data_rows: list[int],
        start_row: int,
    ) -> list[dict]:
        """Validate that required primary key columns are non-empty.

        Args:
            row_validation: The ``row_validation`` dict from the schema.
            data_rows: The list of data row numbers.
            start_row: The data start row (for fallback cell references).

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []

        for pk_key, pk_rule in row_validation.items():
            if pk_key.startswith("_") or not isinstance(pk_rule, dict):
                continue
            if pk_rule.get("required") is not True:
                continue

            pk_column_letter = self._column_map.get(pk_key)
            if not pk_column_letter:
                issues.append(
                    make_issue(
                        "warning",
                        f"B{start_row}",
                        f"Clé primaire '{pk_key}' introuvable dans les en-têtes — "
                        f"validation des lignes de données impossible pour cette clé.",
                        pk_key,
                        "data",
                    )
                )
                continue

            pk_message = pk_rule.get(
                "message", f"Valeur manquante pour '{pk_key}' en ligne {{row}}."
            )
            for row in data_rows:
                value = read_cell_by_row_column(self.worksheet, row, pk_column_letter)
                if value is None or (isinstance(value, str) and not value.strip()):
                    issues.append(
                        make_issue(
                            "error",
                            cell_reference(pk_column_letter, row),
                            pk_message,
                            pk_key,
                            "data",
                        )
                    )

        return issues

    def _check_column_enum(
        self,
        column_entry: dict,
        data_rows: list[int],
        issues: list[dict],
        module_key: str,
    ) -> None:
        """Check enum constraints on data row cells for a header column.

        Handles fixed columns (flat and conditional enums) and recurses
        into group sub_columns.

        Args:
            column_entry: The column entry dict from the schema.
            data_rows: The list of data row numbers.
            issues: The issues list to append to.
            module_key: The module key for issue attribution.
        """
        position = column_entry.get("position", "fixed")

        if position == "fixed":
            validation = column_entry.get("validation", {})
            enum_values = validation.get("enum")
            conditional_enum = validation.get("conditional_enum")

            if not enum_values and not conditional_enum:
                return

            column_letter = column_entry.get("column")
            if not column_letter:
                canonical = column_entry.get("canonical_name", "")
                column_letter = self._column_map.get(canonical)
            if not column_letter:
                return

            header_name = column_entry.get(
                "expected", column_entry.get("canonical_name", "")
            )

            if conditional_enum:
                self._check_conditional_enum(
                    column_letter, header_name, conditional_enum,
                    data_rows, issues, module_key,
                )
            else:
                self._check_flat_enum(
                    column_letter, header_name, enum_values,
                    data_rows, issues, module_key,
                )

        elif position == "group":
            for sub_column in column_entry.get("sub_columns", []):
                self._check_column_enum(sub_column, data_rows, issues, module_key)

    def _check_flat_enum(
        self,
        column_letter: str,
        header_name: str,
        enum_values: list[str],
        data_rows: list[int],
        issues: list[dict],
        module_key: str,
    ) -> None:
        """Check a flat enum constraint on data rows.

        Args:
            column_letter: The column letter to check.
            header_name: The display name for issue messages.
            enum_values: The list of allowed values.
            data_rows: The list of data row numbers.
            issues: The issues list to append to.
            module_key: The module key for issue attribution.
        """
        for row in data_rows:
            value = read_cell_by_row_column(self.worksheet, row, column_letter)
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            string_value = str(value).strip()
            if string_value not in enum_values:
                issues.append(
                    make_issue(
                        "error",
                        cell_reference(column_letter, row),
                        f"Ligne {row}, colonne '{header_name}' ({column_letter}): "
                        f"'{string_value}' n'est pas une valeur autorisée. "
                        f"Valeurs acceptées : {enum_values}.",
                        module_key,
                        "data",
                    )
                )

    def _check_conditional_enum(
        self,
        column_letter: str,
        header_name: str,
        conditional_enum: dict,
        data_rows: list[int],
        issues: list[dict],
        module_key: str,
    ) -> None:
        """Check a conditional enum constraint on data rows.

        The allowed values depend on another column's value.

        Args:
            column_letter: The column letter to check.
            header_name: The display name for issue messages.
            conditional_enum: A dict with ``depends_on_column`` and
                ``values_by_parent``.
            data_rows: The list of data row numbers.
            issues: The issues list to append to.
            module_key: The module key for issue attribution.
        """
        parent_column = conditional_enum.get("depends_on_column")
        values_by_parent = conditional_enum.get("values_by_parent", {})

        for row in data_rows:
            value = read_cell_by_row_column(self.worksheet, row, column_letter)
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            string_value = str(value).strip()
            parent_value = read_cell_by_row_column(self.worksheet, row, parent_column)
            parent_string = str(parent_value).strip() if parent_value else ""
            allowed = values_by_parent.get(parent_string)

            if allowed is None:
                issues.append(
                    make_issue(
                        "warning",
                        cell_reference(column_letter, row),
                        f"Ligne {row}, colonne '{header_name}' ({column_letter}): "
                        f"la valeur parente '{parent_string}' (col {parent_column}) "
                        f"n'a pas de liste de valeurs conditionnelles connue.",
                        module_key,
                        "data",
                    )
                )
            elif string_value not in allowed:
                issues.append(
                    make_issue(
                        "error",
                        cell_reference(column_letter, row),
                        f"Ligne {row}, colonne '{header_name}' ({column_letter}): "
                        f"'{string_value}' n'est pas autorisé quand "
                        f"'{parent_string}' est sélectionné (col {parent_column}). "
                        f"Valeurs acceptées : {allowed}.",
                        module_key,
                        "data",
                    )
                )

    def _validate_column_json_logic(self, data_rows: list[int]) -> list[dict]:
        """Validate column-level json_logic rules across all data rows.

        Args:
            data_rows: The list of data row numbers.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []

        for module_key, module_definition in self.header_modules.items():
            if module_key.startswith("_") or not isinstance(module_definition, dict):
                continue
            for column_entry in self._iter_all_leaf_columns(module_definition):
                column_validation = column_entry.get("validation") or {}
                json_logic_definition = column_validation.get("json_logic")
                if not json_logic_definition:
                    continue

                canonical = column_entry.get("canonical_name", "")
                column_letter = (
                    column_entry.get("column") or self._column_map.get(canonical)
                )
                if not column_letter:
                    continue

                entries = (
                    json_logic_definition
                    if isinstance(json_logic_definition, list)
                    else [json_logic_definition]
                )
                combine_operator = column_validation.get("json_logic_op", "and")

                for row in data_rows:
                    row_data = {
                        cn: read_cell_by_row_column(self.worksheet, row, cl)
                        for cn, cl in self._column_map.items()
                        if cl
                    }
                    # Override with this specific column for duplicate canonicals
                    row_data[canonical] = read_cell_by_row_column(
                        self.worksheet, row, column_letter
                    )
                    try:
                        results = [
                            evaluate_json_logic(entry.get("rule", entry), row_data)
                            for entry in entries
                        ]
                        combined = (
                            any(results)
                            if combine_operator == "or"
                            else all(results)
                        )
                        if not combined:
                            failed = [
                                entry.get("description", "")
                                for entry, result in zip(entries, results)
                                if not result and entry.get("description")
                            ]
                            detail = (
                                "; ".join(failed) if failed else "règle non satisfaite"
                            )
                            issues.append(
                                make_issue(
                                    "error",
                                    cell_reference(column_letter, row),
                                    f"Ligne {row}, colonne "
                                    f"'{column_entry.get('expected', canonical)}' "
                                    f"({column_letter}): {detail}",
                                    module_key,
                                    "data",
                                )
                            )
                    except Exception as exc:
                        issues.append(
                            make_issue(
                                "warning",
                                cell_reference(column_letter, row),
                                f"Ligne {row}: Erreur JsonLogic colonne "
                                f"'{canonical}': {exc}",
                                module_key,
                                "data",
                            )
                        )

        return issues

    def _validate_module_json_logic(self, data_rows: list[int]) -> list[dict]:
        """Validate module-level json_logic rules across all data rows.

        Args:
            data_rows: The list of data row numbers.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []

        for module_key, module_definition in self.header_modules.items():
            if module_key.startswith("_") or not isinstance(module_definition, dict):
                continue
            module_validation = module_definition.get("validation", {})
            module_json_logic = module_validation.get("json_logic")
            if not module_json_logic:
                continue

            entries = (
                module_json_logic
                if isinstance(module_json_logic, list)
                else [module_json_logic]
            )
            combine_operator = module_validation.get("json_logic_op", "and")
            module_cols = self._module_columns.get(module_key, [])
            start_column = module_cols[0] if module_cols else None

            for row in data_rows:
                row_data = {
                    cn: read_cell_by_row_column(self.worksheet, row, cl)
                    for cn, cl in self._column_map.items()
                    if cl
                }
                try:
                    results = [
                        evaluate_json_logic(entry.get("rule", entry), row_data)
                        for entry in entries
                    ]
                    combined = (
                        any(results) if combine_operator == "or" else all(results)
                    )
                    if not combined:
                        failed = [
                            entry.get("description", "")
                            for entry, result in zip(entries, results)
                            if not result and entry.get("description")
                        ]
                        detail = (
                            "; ".join(failed)
                            if failed
                            else "règle de module non satisfaite"
                        )
                        issues.append(
                            make_issue(
                                "error",
                                cell_reference(start_column or "A", row),
                                f"Ligne {row}, module '{module_key}': {detail}",
                                module_key,
                                "data",
                            )
                        )
                except Exception as exc:
                    issues.append(
                        make_issue(
                            "warning",
                            cell_reference(start_column or "A", row),
                            f"Ligne {row}: Erreur JsonLogic module "
                            f"'{module_key}': {exc}",
                            module_key,
                            "data",
                        )
                    )

        return issues

    # -- Sections-based data validation -------------------------------------

    def _validate_sections_data(self) -> list[dict]:
        """Validate data rows for sections-based schemas.

        Checks required columns and type constraints.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []

        for section in self.sections:
            if section.get("layout") != "records":
                continue

            start_row = section.get("data", {}).get("start_row", 2)
            end_rule = section.get("data", {}).get("end_rule", "first_empty_row")
            primary_key = section.get("headers", {}).get("primary_key")

            data_rows: list[int] = []
            for row in range(start_row, self.worksheet.max_row + 1):
                if end_rule == "first_empty_row":
                    from cell_helpers import is_row_empty
                    if is_row_empty(self.worksheet, row):
                        break
                data_rows.append(row)

            if not data_rows:
                issues.append(
                    make_issue(
                        "warning",
                        f"A{start_row}",
                        f"Aucune ligne de données trouvée à partir de "
                        f"la ligne {start_row}.",
                        section.get("key", ""),
                        "data",
                    )
                )
                continue

            for column_definition in section.get("headers", {}).get("columns", []):
                column_type = column_definition.get("type", "single")
                if column_type == "single":
                    issues.extend(
                        self._validate_data_column(
                            column_definition, data_rows, primary_key
                        )
                    )
                elif column_type == "group":
                    for sub_column in column_definition.get("sub_columns", []):
                        issues.extend(
                            self._validate_data_column(
                                sub_column, data_rows, primary_key
                            )
                        )

        return issues

    def _validate_data_column(
        self,
        column_definition: dict,
        data_rows: list[int],
        primary_key: str | None,
    ) -> list[dict]:
        """Validate data cells for a single column definition.

        Checks required/primary key emptiness and type constraints.

        Args:
            column_definition: The column definition dict.
            data_rows: The list of data row numbers.
            primary_key: The canonical name of the primary key column.

        Returns:
            A list of issue dicts.
        """
        issues: list[dict] = []
        column_letter = column_definition.get("column")
        canonical = column_definition.get("canonical_name", "")
        required = column_definition.get("required", False)
        value_type = column_definition.get("value_type", "string")
        is_primary_key = canonical == primary_key

        if not column_letter:
            return issues

        for row in data_rows:
            value = read_cell_by_row_column(self.worksheet, row, column_letter)
            is_empty = value is None or (isinstance(value, str) and not value.strip())

            if is_empty:
                if is_primary_key or required:
                    issues.append(
                        make_issue(
                            "error",
                            cell_reference(column_letter, row),
                            f"Valeur manquante en {column_letter}{row} "
                            f"(colonne {'clé primaire ' if is_primary_key else 'requise '}"
                            f"'{column_definition.get('expected', canonical)}').",
                            canonical,
                            "data",
                        )
                    )
                continue

            # Type validation
            if value_type == "number" and not isinstance(value, (int, float)):
                try:
                    float(value)
                except (ValueError, TypeError):
                    issues.append(
                        make_issue(
                            "warning",
                            cell_reference(column_letter, row),
                            f"Valeur en {column_letter}{row} ('{value}') "
                            f"devrait être un nombre.",
                            canonical,
                            "data",
                        )
                    )
            elif value_type == "percentage" and not isinstance(value, (int, float)):
                try:
                    float(str(value).rstrip("%"))
                except (ValueError, TypeError):
                    issues.append(
                        make_issue(
                            "warning",
                            cell_reference(column_letter, row),
                            f"Valeur en {column_letter}{row} ('{value}') "
                            f"devrait être un pourcentage.",
                            canonical,
                            "data",
                        )
                    )

        return issues

    # -- Helpers ------------------------------------------------------------

    @staticmethod
    def _iter_all_leaf_columns(module_definition: dict):
        """Yield all leaf (fixed) column defs from a module, including group sub_columns.

        Args:
            module_definition: A header module definition dict.

        Yields:
            Individual column entry dicts with ``position: fixed``.
        """
        for column in module_definition.get("columns", []):
            position = column.get("position", "fixed")
            if position == "fixed":
                yield column
            elif position == "group":
                for sub_column in column.get("sub_columns", []):
                    if sub_column.get("position") == "fixed":
                        yield sub_column


# ---------------------------------------------------------------------------
# Helpers shared across validators
# ---------------------------------------------------------------------------


def _is_date_string(value) -> bool:
    """Check whether a value looks like an ISO date string.

    Args:
        value: The value to check.

    Returns:
        True if the value is a datetime or parses as an ISO date.
    """
    if isinstance(value, datetime):
        return True
    try:
        datetime.fromisoformat(str(value))
        return True
    except (ValueError, TypeError):
        return False
