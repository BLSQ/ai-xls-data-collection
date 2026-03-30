"""Excel workbook reader for structure analysis.

Loads the workbook in two modes (values-only and full with formulas/styles),
provides sheet selection and builds the rich text representation sent to the LLM.
"""

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

MAX_ANALYSIS_ROWS = 35
MAX_ANALYSIS_COLS = 60

# Sheet names to skip when auto-selecting (case-insensitive).
_SKIP_SHEETS = {"guide", "liste des catégories", "liste des categories", "observations"}


class ExcelReader:
    """Reads an Excel workbook for AI-based structure analysis.

    Opens the file twice:
    - ``workbook_values``  (data_only=True)  — resolved cell values and styles.
    - ``workbook_full``    (data_only=False) — formulas, data-validations.

    Both handles are kept alive until :meth:`close` is called so that
    downstream code can read from either representation.
    """

    def __init__(self, excel_path: str | Path):
        """Open the workbook in both value and formula modes.

        Args:
            excel_path: Path to the ``.xlsx`` file on disk.
        """
        self._path = Path(excel_path)
        self.workbook_values = openpyxl.load_workbook(str(self._path), data_only=True)
        self.workbook_full = openpyxl.load_workbook(str(self._path), data_only=False)

    # ------------------------------------------------------------------
    # Sheet selection
    # ------------------------------------------------------------------

    def select_sheets(self, sheet_name: str | None = None):
        """Return the target worksheet from both workbook modes.

        If *sheet_name* is given, that sheet is selected directly.
        Otherwise the first sheet whose name is **not** in the skip list
        (guide, observations, …) is chosen automatically.

        Returns:
            A tuple ``(worksheet_values, worksheet_full)`` for the same
            sheet in both workbook modes.
        """
        if sheet_name:
            return (
                self.workbook_values[sheet_name],
                self.workbook_full[sheet_name],
            )

        for name in self.workbook_values.sheetnames:
            if name.strip().lower() not in _SKIP_SHEETS:
                return (
                    self.workbook_values[name],
                    self.workbook_full[name],
                )

        first = self.workbook_values.sheetnames[0]
        return (
            self.workbook_values[first],
            self.workbook_full[first],
        )

    # ------------------------------------------------------------------
    # Text representation for the LLM prompt
    # ------------------------------------------------------------------

    def build_text_representation(
        self,
        worksheet_values,
        worksheet_full,
        max_rows: int = MAX_ANALYSIS_ROWS,
    ) -> str:
        """Build a rich text dump of the sheet for the LLM prompt.

        Includes cell values, styles (bold / background colour), merged-cell
        ranges and data-validation rules — everything an LLM needs to reason
        about the spreadsheet layout without seeing a pixel.

        Args:
            worksheet_values: Worksheet opened with ``data_only=True``.
            worksheet_full:   Worksheet opened with ``data_only=False``.
            max_rows:         Number of structural rows to include.

        Returns:
            Multi-line string ready to be inserted into an LLM prompt.
        """
        max_col = min(worksheet_values.max_column or 1, MAX_ANALYSIS_COLS)
        lines: list[str] = []

        # -- overview --
        lines.append(f"SHEET: '{worksheet_values.title}'")
        lines.append(
            f"Dimensions: {worksheet_values.max_row} rows x {worksheet_values.max_column} columns "
            f"(A-{get_column_letter(worksheet_values.max_column or 1)})"
        )
        lines.append("")

        # -- merged cells --
        merges = sorted(
            worksheet_values.merged_cells.ranges,
            key=lambda m: (m.min_row, m.min_col),
        )
        lines.append(f"MERGED CELLS ({len(merges)} total):")
        for merged_range in merges[:80]:
            value = worksheet_values.cell(merged_range.min_row, merged_range.min_col).value
            value_str = f' = "{str(value)[:200]}"' if value is not None else ""
            lines.append(f"  {merged_range}{value_str}")
        if len(merges) > 80:
            lines.append(f"  ... ({len(merges) - 80} more)")
        lines.append("")

        # -- build merge helpers --
        merge_skip: set[tuple[int, int]] = set()
        merge_map: dict[tuple[int, int], object] = {}
        for merged_range in merges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != (merged_range.min_row, merged_range.min_col):
                        merge_skip.add((row, col))
                    else:
                        merge_map[(row, col)] = merged_range

        # -- cell grid --
        lines.append(
            f"CELL DATA (rows 1-{max_rows}, columns A-{get_column_letter(max_col)}):"
        )
        lines.append('Format: CellRef="value" [style annotations] [merge info]')
        lines.append("")

        for row in range(1, max_rows + 1):
            parts: list[str] = []
            for col in range(1, max_col + 1):
                if (row, col) in merge_skip:
                    continue
                cell = worksheet_values.cell(row, col)
                if cell.value is None:
                    continue

                cell_ref = f"{get_column_letter(col)}{row}"
                value_text = str(cell.value)[:200]

                annotations: list[str] = []
                if cell.font and cell.font.bold:
                    annotations.append("bold")
                background_hex = self._cell_background_hex(cell)
                if background_hex:
                    annotations.append(f"bg:{background_hex}")
                    foreground_hex = self._cell_foreground_hex(cell)
                    if foreground_hex:
                        annotations.append(f"fg:{foreground_hex}")
                if (row, col) in merge_map:
                    annotations.append(f"merged:{merge_map[(row, col)]}")

                annotation_str = f" [{', '.join(annotations)}]" if annotations else ""
                parts.append(f'{cell_ref}="{value_text}"{annotation_str}')

            if parts:
                lines.append(f"  Row {row}: {' | '.join(parts)}")
            else:
                lines.append(f"  Row {row}: (empty)")

        # -- sample data rows (next 5 after structural zone) --
        lines.append("")
        lines.append("SAMPLE DATA ROWS (next 5 rows after structural zone):")
        for row in range(max_rows + 1, min(max_rows + 6, (worksheet_values.max_row or 0) + 1)):
            parts = []
            for col in range(1, max_col + 1):
                if (row, col) in merge_skip:
                    continue
                cell = worksheet_values.cell(row, col)
                if cell.value is None:
                    continue
                cell_ref = f"{get_column_letter(col)}{row}"
                parts.append(f'{cell_ref}="{str(cell.value)[:200]}"')
            if parts:
                lines.append(f"  Row {row}: {' | '.join(parts)}")

        # -- data validations --
        lines.append("")
        lines.append("DATA VALIDATIONS (dropdown lists):")
        if hasattr(worksheet_full, "data_validations") and worksheet_full.data_validations:
            for data_validation in worksheet_full.data_validations.dataValidation:
                if data_validation.type == "list":
                    lines.append(
                        f"  {data_validation.sqref}: formula={data_validation.formula1}"
                    )
        else:
            lines.append("  (none)")

        # -- column widths --
        lines.append("")
        lines.append("COLUMN WIDTHS (non-default):")
        for col_index in range(1, max_col + 1):
            letter = get_column_letter(col_index)
            dimension = worksheet_values.column_dimensions.get(letter)
            if dimension and dimension.width and dimension.width > 12:
                lines.append(f"  {letter}: {dimension.width:.1f}")

        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Colour extraction helpers (static)
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_hex_color(color_object) -> str | None:
        """Safely extract ``#RRGGBB`` from an openpyxl ``Color`` object.

        Returns ``None`` for theme/indexed colours, transparent fills,
        or missing colour data.
        """
        if color_object is None:
            return None
        if getattr(color_object, "type", None) not in (None, "rgb"):
            return None
        rgb = getattr(color_object, "rgb", None)
        if rgb is None or not isinstance(rgb, str):
            return None
        rgb = rgb.strip()
        if not all(c in "0123456789abcdefABCDEF" for c in rgb):
            return None
        if len(rgb) == 8 and rgb[:2] == "00":
            return None  # fully transparent
        if len(rgb) >= 6:
            return f"#{rgb[-6:]}"
        return None

    @staticmethod
    def _cell_background_hex(cell) -> str | None:
        """Extract a cell's background colour as ``#RRGGBB`` or ``None``."""
        if not cell.fill:
            return None
        return ExcelReader._extract_hex_color(cell.fill.fgColor)

    @staticmethod
    def _cell_foreground_hex(cell) -> str | None:
        """Extract a cell's font colour as ``#RRGGBB`` or ``None``."""
        if not cell.font:
            return None
        return ExcelReader._extract_hex_color(cell.font.color)

    # ------------------------------------------------------------------
    # Resource management
    # ------------------------------------------------------------------

    def close(self):
        """Close both workbook handles and release resources."""
        self.workbook_values.close()
        self.workbook_full.close()
