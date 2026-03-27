"""OpenHEXA Pipeline: Excel Structure Proposal via Gemini AI.

Analyzes an Excel template's visual and structural layout using Gemini's
multimodal capabilities.  Produces a structure_proposal.json describing
every section found: metadata key-value pairs, grouped key-value tables,
and record-based data tables.

The structure proposal is designed to be reviewed/edited by a human
(or a webapp) before being consumed by a deterministic schema generator.
"""

import base64
import io
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openhexa.sdk import (
    CustomConnection,
    File,
    current_run,
    parameter,
    pipeline,
    workspace,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models"
DEFAULT_MODEL = "gemini-2.5-flash"
MAX_ANALYSIS_ROWS = 35  # rows sent to the LLM (structural zone)
MAX_ANALYSIS_COLS = 60  # columns sent to the LLM
OUTPUT_FILENAME = "structure_proposal.json"


# ---------------------------------------------------------------------------
# 1. Excel  ->  structured text representation
# ---------------------------------------------------------------------------


def build_text_representation(ws, ws_full, max_rows=MAX_ANALYSIS_ROWS):
    """Build a rich text dump of the sheet for the LLM prompt.

    Includes cell values, styles (bold / background colour), merged-cell
    ranges and data-validation rules — everything an LLM needs to reason
    about the spreadsheet layout without seeing a pixel.
    """
    max_col = min(ws.max_column or 1, MAX_ANALYSIS_COLS)
    lines = []

    # -- overview --
    lines.append(f"SHEET: '{ws.title}'")
    lines.append(
        f"Dimensions: {ws.max_row} rows x {ws.max_column} columns "
        f"(A-{get_column_letter(ws.max_column or 1)})"
    )
    lines.append("")

    # -- merged cells --
    merges = sorted(ws.merged_cells.ranges, key=lambda m: (m.min_row, m.min_col))
    lines.append(f"MERGED CELLS ({len(merges)} total):")
    for mr in merges[:80]:
        val = ws.cell(mr.min_row, mr.min_col).value
        val_str = f' = "{str(val)[:200]}"' if val is not None else ""
        lines.append(f"  {mr}{val_str}")
    if len(merges) > 80:
        lines.append(f"  ... ({len(merges) - 80} more)")
    lines.append("")

    # -- build merge helpers --
    merge_skip = set()
    merge_map = {}  # (row, col) -> MergedCellRange
    for mr in merges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    merge_skip.add((r, c))
                else:
                    merge_map[(r, c)] = mr

    # -- cell grid --
    lines.append(f"CELL DATA (rows 1-{max_rows}, columns A-{get_column_letter(max_col)}):")
    lines.append('Format: CellRef="value" [style annotations] [merge info]')
    lines.append("")

    for row in range(1, max_rows + 1):
        parts = []
        for col in range(1, max_col + 1):
            if (row, col) in merge_skip:
                continue
            cell = ws.cell(row, col)
            if cell.value is None:
                continue

            ref = f"{get_column_letter(col)}{row}"
            val = str(cell.value)[:200]

            annotations = []
            if cell.font and cell.font.bold:
                annotations.append("bold")
            bg = _cell_bg_hex(cell)
            if bg:
                annotations.append(f"bg:{bg}")
                fg_hex = _cell_fg_hex(cell)
                if fg_hex:
                    annotations.append(f"fg:{fg_hex}")
            if (row, col) in merge_map:
                annotations.append(f"merged:{merge_map[(row, col)]}")

            ann = f" [{', '.join(annotations)}]" if annotations else ""
            parts.append(f'{ref}="{val}"{ann}')

        if parts:
            lines.append(f"  Row {row}: {' | '.join(parts)}")
        else:
            lines.append(f"  Row {row}: (empty)")

    # -- sample data rows (rows max_rows+1 .. max_rows+5) for record detection --
    lines.append("")
    lines.append("SAMPLE DATA ROWS (next 5 rows after structural zone):")
    for row in range(max_rows + 1, min(max_rows + 6, (ws.max_row or 0) + 1)):
        parts = []
        for col in range(1, max_col + 1):
            if (row, col) in merge_skip:
                continue
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            ref = f"{get_column_letter(col)}{row}"
            parts.append(f'{ref}="{str(cell.value)[:200]}"')
        if parts:
            lines.append(f"  Row {row}: {' | '.join(parts)}")

    # -- data validations --
    lines.append("")
    lines.append("DATA VALIDATIONS (dropdown lists):")
    if hasattr(ws_full, "data_validations") and ws_full.data_validations:
        for dv in ws_full.data_validations.dataValidation:
            if dv.type == "list":
                lines.append(f"  {dv.sqref}: formula={dv.formula1}")
    else:
        lines.append("  (none)")

    # -- column widths (hint for the LLM) --
    lines.append("")
    lines.append("COLUMN WIDTHS (non-default):")
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width and dim.width > 12:
            lines.append(f"  {letter}: {dim.width:.1f}")

    return "\n".join(lines)


def _extract_hex_color(color_obj):
    """Safely extract #RRGGBB from an openpyxl Color object, or None.

    openpyxl Color.rgb can be a hex string ('FF1F4E79'), a type descriptor,
    or None.  Theme/indexed colours don't have usable rgb values.
    """
    if color_obj is None:
        return None
    # Only process rgb-type colours (skip theme, indexed, auto)
    if getattr(color_obj, "type", None) not in (None, "rgb"):
        return None
    rgb = getattr(color_obj, "rgb", None)
    if rgb is None or not isinstance(rgb, str):
        return None
    # Must look like a hex string: 6 or 8 hex chars
    rgb = rgb.strip()
    if not all(c in "0123456789abcdefABCDEF" for c in rgb):
        return None
    if len(rgb) == 8 and rgb[:2] == "00":
        return None  # fully transparent
    if len(rgb) >= 6:
        return f"#{rgb[-6:]}"
    return None


def _cell_bg_hex(cell):
    """Extract background colour as #RRGGBB or None."""
    if not cell.fill:
        return None
    return _extract_hex_color(cell.fill.fgColor)


def _cell_fg_hex(cell):
    """Extract font colour as #RRGGBB or None."""
    if not cell.font:
        return None
    return _extract_hex_color(cell.font.color)


# ---------------------------------------------------------------------------
# 2. Excel  ->  PIL image (multimodal input for Gemini)
# ---------------------------------------------------------------------------


def render_sheet_image(ws, max_rows=MAX_ANALYSIS_ROWS, max_cols=MAX_ANALYSIS_COLS):
    """Render the top-left region of a worksheet as a PNG image.

    Returns the PNG bytes, or None if Pillow is unavailable.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None

    COL_W = 90
    ROW_H = 20
    HDR_H = 22
    LBL_W = 35

    actual_cols = min(ws.max_column or 1, max_cols)
    actual_rows = min(ws.max_row or 1, max_rows)

    width = LBL_W + actual_cols * COL_W + 2
    height = HDR_H + actual_rows * ROW_H + 2

    img = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    # -- fonts --
    font, font_b = _load_fonts()

    # -- merge helpers --
    skip_cells = set()
    merge_spans = {}  # (row, col) -> (rowspan, colspan)
    for mr in ws.merged_cells.ranges:
        merge_spans[(mr.min_row, mr.min_col)] = (
            mr.max_row - mr.min_row + 1,
            mr.max_col - mr.min_col + 1,
        )
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    skip_cells.add((r, c))

    # -- column headers --
    for c in range(1, actual_cols + 1):
        x = LBL_W + (c - 1) * COL_W
        draw.rectangle([x, 0, x + COL_W, HDR_H], fill="#E8E8E8", outline="#CCCCCC")
        draw.text((x + 4, 4), get_column_letter(c), fill="#333333", font=font_b)

    # -- cells --
    for row in range(1, actual_rows + 1):
        y = HDR_H + (row - 1) * ROW_H

        # row label
        draw.rectangle([0, y, LBL_W, y + ROW_H], fill="#E8E8E8", outline="#CCCCCC")
        draw.text((3, y + 3), str(row), fill="#555555", font=font)

        for col in range(1, actual_cols + 1):
            if (row, col) in skip_cells:
                continue

            x = LBL_W + (col - 1) * COL_W

            # cell dimensions (handle merges)
            if (row, col) in merge_spans:
                rs, cs = merge_spans[(row, col)]
                cw = min(cs, actual_cols - col + 1) * COL_W
                ch = min(rs, actual_rows - row + 1) * ROW_H
            else:
                cw = COL_W
                ch = ROW_H

            cell = ws.cell(row, col)
            bg = _cell_bg_hex(cell) or "#FFFFFF"
            fg = "#000000"

            # auto-contrast for dark backgrounds
            try:
                r_val = int(bg[1:3], 16)
                g_val = int(bg[3:5], 16)
                b_val = int(bg[5:7], 16)
                if (r_val + g_val + b_val) / 3 < 128:
                    fg = "#FFFFFF"
            except (ValueError, IndexError):
                pass

            draw.rectangle([x, y, x + cw, y + ch], fill=bg, outline="#CCCCCC")

            if cell.value is not None:
                text = str(cell.value)
                max_chars = max((cw - 8) // 6, 4)
                if len(text) > max_chars:
                    text = text[: max_chars - 2] + ".."
                use_font = font_b if (cell.font and cell.font.bold) else font
                draw.text((x + 4, y + 3), text, fill=fg, font=use_font)

    # -- encode as PNG bytes --
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _load_fonts():
    """Try to load DejaVu Sans; fall back to PIL default."""
    from PIL import ImageFont

    paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
    ]
    for p in paths:
        try:
            return (
                ImageFont.truetype(p, 10),
                ImageFont.truetype(p.replace("Sans.", "Sans-Bold."), 10),
            )
        except OSError:
            continue
    default = ImageFont.load_default()
    return default, default


# ---------------------------------------------------------------------------
# 3. Gemini API caller
# ---------------------------------------------------------------------------


def call_gemini(prompt, image_bytes, api_key, model=DEFAULT_MODEL):
    """Call Gemini with a multimodal prompt (text + optional image).

    Returns the raw text response from the model.
    """
    url = f"{GEMINI_API_URL}/{model}:generateContent"

    parts = [{"text": prompt}]
    if image_bytes:
        parts.append(
            {
                "inline_data": {
                    "mime_type": "image/png",
                    "data": base64.b64encode(image_bytes).decode("ascii"),
                }
            }
        )

    body = {
        "contents": [{"parts": parts}],
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": 65536,
        },
    }

    current_run.log_info(
        f"Calling Gemini ({model}) — prompt ~{len(prompt)} chars"
        + (f", image {len(image_bytes)} bytes" if image_bytes else "")
    )

    resp = requests.post(url, params={"key": api_key}, json=body, timeout=120)
    resp.raise_for_status()
    data = resp.json()

    # Extract text from response
    try:
        candidates = data["candidates"]
        text = candidates[0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError) as exc:
        current_run.log_warning(f"Unexpected Gemini response shape: {json.dumps(data)[:500]}")
        raise RuntimeError(f"Could not parse Gemini response: {exc}") from exc

    # Log token usage if available
    usage = data.get("usageMetadata", {})
    if usage:
        current_run.log_info(
            f"Gemini tokens — prompt: {usage.get('promptTokenCount', '?')}, "
            f"response: {usage.get('candidatesTokenCount', '?')}, "
            f"total: {usage.get('totalTokenCount', '?')}"
        )

    return text


# ---------------------------------------------------------------------------
# 4. Response parser
# ---------------------------------------------------------------------------


def parse_structure_json(raw_text):
    """Extract and parse JSON from Gemini's response text.

    Handles both raw JSON and JSON wrapped in ```json ... ``` markers.
    """
    # Try raw JSON first
    text = raw_text.strip()
    if text.startswith("{"):
        return json.loads(text)

    # Try extracting from code block
    m = re.search(r"```(?:json)?\s*\n?(.*?)```", text, re.DOTALL)
    if m:
        return json.loads(m.group(1).strip())

    # Last resort: find first { to last }
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return json.loads(text[start : end + 1])

    raise ValueError(f"Could not extract JSON from response:\n{text[:500]}")


# ---------------------------------------------------------------------------
# 5. Prompt construction
# ---------------------------------------------------------------------------

SCHEMA_FORMAT_SPEC = """\
You are a spreadsheet structure analyst.  Given the cell data, merged cell
ranges, styles, and data validations from an Excel worksheet, identify
every distinct SECTION and classify its layout.

## The Three Layout Patterns

### 1. `key_value`  — Simple label -> value metadata
A cell contains a label (usually bold, often ending with ":"), and the
adjacent cell (to the right, or across a merged range) contains a value.
These are typically found in the upper part of a sheet, above data tables.

Example:
```
Row 3: B="Nom du projet:" [bold] | C="PROSANTE" [merged C3:I3]
Row 4: B="Date début:" [bold]    | C="2023-01-01"
```

### 2. `grouped_key_value`  — Repeating column group (transposed table)
A label cell is followed by N element columns (e.g., "Bailleur 1",
"Bailleur 2", ...), and below them are property rows with a value per
element.  This is a TABLE embedded in metadata, but oriented
column-per-entity rather than row-per-record.

How to recognise:
- A header row with a group label + N numbered/named element columns
- Below it: property rows where each element column has a value
- The number of elements may vary (some may be empty)

Example:
```
Row 8:  B="Bailleur(s)" [bold] | C="Bailleur 1" | D="Bailleur 2" | ...
Row 9:  B="Nom"                | C="USAID"      | D="Gavi"       | ...
Row 10: B="Montant"            | C=500000       | D=300000       | ...
```

### 3. `records`  — Data table with column headers + data rows
One or two header rows define columns, followed by many data rows.
Headers may have GROUPS: a parent header merged across several columns,
with sub-column headers in the row below.
Some columns may be DYNAMIC (variable count, e.g., year columns, province names).

How to recognise:
- Bold / coloured title row(s) spanning the full width -> module titles
- Below: column header row(s) with individual column names
- Below: data rows (numbers, text, dates)
- Data validations (dropdowns) often apply to entire data columns

Example:
```
Row 14: B:I merged "Module 3: Budget data" [bold, colored bg]
Row 15: B="Budget line" | C="Total" [merged C15:D15] | E="Province 1" | ...
Row 16:                 | C="Planned" | D="Actual"   |               | ...
Row 17: B="Vaccines"    | C=50000     | D=45000       | E=30%         | ...
```

## How to identify sections

1. **Title rows**: Bold cells with coloured background spanning multiple
   columns are section/module titles.  They mark section boundaries.
2. **Metadata vs. data**: Above the first record-table title row is
   metadata (key_value or grouped_key_value).  Below it is data (records).
3. **Grouped key_value**: Look for a label + numbered/named element
   columns in the same row, with property rows below.
4. **Records columns** come in two types:
   - `"type": "single"` — a standalone column with one header in one row.
     Produces a flat indicator name: `canonical_name`.
   - `"type": "group"` — a parent header (row 15, merged across columns)
     with sub-column headers (row 16).  Produces concatenated indicator
     names: `parent_canonical_name + "_" + sub_canonical_name`.
   A single module CAN MIX both types.  For example, Module 8 has three
   single columns (NA, National, Central) followed by a group column
   (Provincial → province sub-columns).
   Within a group's sub_columns, each sub is either `"fixed": true`
   (known, stable name) or `"fixed": false` (dynamic — the count and
   names may vary across template versions, e.g., province names).
5. **Data validations**: Dropdowns on data columns indicate enum constraints.

## Important guidelines

- Detect ALL sections, including ones that span only a few rows.
- For `records` sections, carefully identify: which row(s) are headers,
  which row is the first data row, and the column structure.
- `headers.rows` is PER-SECTION.  If a module has ONLY `"type": "single"`
  columns, set `"rows": [15]` (one row).  If the module has ANY
  `"type": "group"` column (parent header + sub-columns), include both
  rows: `"rows": [15, 16]`.  A module CAN mix single and group columns.
- For group headers in records, identify the parent and its sub-columns.
- For dynamic columns (years, provinces, categories), note the pattern.
- Assign a short snake_case `key` to each section and field.
- `canonical_name` for record columns should be snake_case of the header.
- Include the `area` (row and column range) for each section.
- For value types, infer from example values: "string", "number", "date",
  "percentage".
- Note data validations that apply to columns (enum constraints).

An image of the spreadsheet is attached. Use it to:
- Confirm section boundaries where the visual layout shifts (color bands,
  spacing, or density changes between regions).
- Identify grouped columns that share a visual pattern (same background
  color, similar formatting) even if their names don't follow a numbered
  convention.
- Distinguish metadata regions (sparse, label-value pairs) from data
  tables (dense grids) when the text representation is ambiguous.
Do not rely solely on the image — the text representation is authoritative
for exact cell values and references.


## Output format

Produce a single JSON object with this structure:

```json
{
  "version": "2.0.0",
  "source_sheet": "<sheet name>",
  "schema_meta": {
    "fingerprint": {
      "cell": "<cell ref of the first section title>",
      "expected": "<title text>"
    },
    "locale": "<detected language: fr, en, etc.>",
    "skip_sheets": ["<sheet names that look like reference/guide sheets>"]
  },
  "structure_landmarks": {
    "metadata_area": {"start_row": 1, "end_row": 13},
    "title_row": 14,
    "header_rows": [15, 16],
    "data_start_row": 17,
    "last_data_col": "BF"
  },
  "sections": [
    {
      "key": "<snake_case section name>",
      "title": {
        "cell": "<cell ref>",
        "expected": "<title text>"
      },
      "layout": "key_value",
      "area": {"start_row": N, "end_row": N, "start_col": "B", "end_col": "I"},
      "fields": [
        {
          "key": "<snake_case>",
          "label": {"cell": "<ref>", "expected": "<label text>"},
          "value": {
            "cell": "<ref>",
            "type": "string|number|date",
            "required": true|false,
            "validation": {"enum": ["...", "..."]}
          }
        }
      ]
    },
    {
      "key": "<snake_case>",
      "title": {"cell": "<ref>", "expected": "<text>"},
      "layout": "grouped_key_value",
      "area": {"start_row": N, "end_row": N, "start_col": "B", "end_col": "I"},
      "elements": {
        "header_row": N,
        "columns": ["C", "D", "E"],
        "labels": ["Element 1", "Element 2", "..."],
        "count_is_dynamic": true
      },
      "fields": [
        {
          "key": "<snake_case>",
          "label": {"row": N, "col": "B", "expected": "<text>"},
          "type": "string|number|date",
          "required_rule": "always|first_required|if_element_active"
        }
      ]
    },
    {
      "key": "<snake_case>",
      "title": {"cell": "<ref>", "expected": "<module title text>"},
      "layout": "records",
      "area": {"start_row": N, "end_row": "dynamic", "start_col": "B", "end_col": "I"},
      "headers": {
        "rows": [15, 16],
        "primary_key": "<canonical_name of the required column>",
        "columns": [
          {
            "type": "single",
            "column": "B",
            "header_row": 15,
            "expected": "<header text>",
            "canonical_name": "<snake_case>",
            "value_type": "string|number|date|percentage",
            "required": true|false
          },
          {
            "type": "group",
            "parent": {
              "header_row": 15,
              "expected": "<parent header text>",
              "canonical_name": "<snake_case of group>",
              "start_col": "C",
              "end_col": "F"
            },
            "sub_columns": [
              {
                "fixed": true,
                "column": "C",
                "header_row": 16,
                "expected": "<sub-header text>",
                "canonical_name": "<snake_case>",
                "value_type": "number"
              },
              {
                "fixed": false,
                "header_row": 16,
                "pattern": "regex to match dynamic headers",
                "canonical_prefix": "<prefix>",
                "value_type": "number",
                "description": "what these dynamic columns represent"
              }
            ]
          }
        ]
      },
      "data": {
        "start_row": 17,
        "end_rule": "first_empty_row"
      },
    }
  ],
  "data_validations_summary": [
    {
      "cells": "<sqref>",
      "type": "list",
      "formula": "<formula1 value>",
      "resolved_values": ["val1", "val2"]
    }
  ],
  "validation_config": {
    "label_comparison": {
      "case_sensitive": false,
      "normalize_accents": true,
      "strip_whitespace": true
    }
  }
}
```

## Key fields explained

- **`structure_landmarks`**: Global row/column boundaries. `metadata_area`
  covers key_value and grouped_key_value sections. `title_row` is the row
  where record module titles are (bold colored cells spanning columns).
  `header_rows` are the column header row(s) for data tables.
  `data_start_row` is where records begin. `last_data_col` is the
  rightmost column with actual data.

- **`schema_meta.fingerprint`**: The cell + text used to verify whether a
  submitted workbook matches this template (usually the first module title).

- **`data_validations_summary`**: List the validations from the DATA
  VALIDATIONS section. Set `resolved_values` to null if you cannot
  determine the actual values (named ranges, INDIRECT formulas) — the
  deterministic pipeline will resolve them later.

- **`headers.primary_key`**: The canonical_name of the column that must
  not be empty for a data row to be valid (usually the first/leftmost
  column — the "line item" identifier).

- **Column `type` (single vs group)**: Each column in `headers.columns`
  has `"type": "single"` or `"type": "group"`.  A module CAN mix both.
  - `single`: one header cell, one column.  Indicator = `canonical_name`.
  - `group`: merged parent header spanning N sub-columns.  Each sub has
    `"fixed": true` (stable name) or `"fixed": false` (dynamic — name/count
    varies across template versions).  Indicator = `parent_canonical + "_" + sub_canonical`.
  Example: Module 8 has single columns (NA, National, Central) then a
  group (Provincial → 26 province sub-columns marked `"fixed": false`).

Now analyse the following spreadsheet and produce the structure_proposal JSON.
"""


def build_prompt(text_repr, sheet_name, user_guidelines=None):
    """Assemble the full prompt: format spec + user guidelines + sheet data."""
    parts = [SCHEMA_FORMAT_SPEC, ""]

    if user_guidelines and user_guidelines.strip():
        parts.append("=== USER GUIDELINES ===")
        parts.append(
            "The user provided the following domain-specific instructions. "
            "Use them to improve your analysis — they take priority over "
            "your own heuristics when there is a conflict."
        )
        parts.append("")
        parts.append(user_guidelines.strip())
        parts.append("")

    parts.append(f"=== SPREADSHEET DATA FOR SHEET '{sheet_name}' ===")
    parts.append("")
    parts.append(text_repr)
    parts.append("")
    parts.append("=== END OF DATA ===")
    parts.append("")
    parts.append(
        "Produce the structure_proposal JSON for this sheet.  "
        "Be thorough — identify every section, every field, every column."
    )

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# 6. Data validation pre-resolution (deterministic, complements LLM)
# ---------------------------------------------------------------------------


def extract_resolved_validations(ws_full, wb_full):
    """Extract and resolve data validations deterministically.

    Returns a list of dicts with resolved dropdown values where possible.
    This supplements the LLM's analysis with precise values.
    """
    results = []
    if not hasattr(ws_full, "data_validations") or ws_full.data_validations is None:
        return results

    for dv in ws_full.data_validations.dataValidation:
        if dv.type != "list":
            continue

        formula = str(dv.formula1 or "")
        values = _try_resolve_formula(ws_full, wb_full, formula)

        results.append(
            {
                "cells": str(dv.sqref),
                "type": "list",
                "formula": formula,
                "resolved_values": values if values else None,
                "allow_blank": bool(dv.allow_blank),
            }
        )

    return results


def _try_resolve_formula(ws, wb, formula_str):
    """Best-effort resolution of a data validation formula."""
    formula_str = formula_str.strip('"')
    if not formula_str:
        return None

    # INDIRECT — skip (would need cell values to resolve)
    if formula_str.upper().startswith("INDIRECT("):
        return None

    # Cell range reference
    if "$" in formula_str or ":" in formula_str or "!" in formula_str:
        return _resolve_range(ws, wb, formula_str)

    # Named range
    if hasattr(wb, "defined_names"):
        defn = wb.defined_names.get(formula_str)
        if defn:
            return _resolve_named_range(wb, defn)

    # Inline list (semicolons or commas)
    if ";" in formula_str:
        return [v.strip() for v in formula_str.split(";") if v.strip()]
    if "," in formula_str:
        return [v.strip() for v in formula_str.split(",") if v.strip()]

    return [formula_str] if formula_str else None


def _resolve_range(ws, wb, formula_str):
    """Resolve a cell range reference to values."""
    try:
        if "!" in formula_str:
            sheet_part, range_part = formula_str.rsplit("!", 1)
            sheet_name = sheet_part.strip("'\"")
            if sheet_name not in wb.sheetnames:
                return None
            source = wb[sheet_name]
        else:
            source = ws
            range_part = formula_str
        range_part = range_part.replace("$", "")
        values = []
        for row_or_cell in source[range_part]:
            cells = row_or_cell if isinstance(row_or_cell, tuple) else (row_or_cell,)
            for cell in cells:
                if cell.value is not None and str(cell.value).strip():
                    values.append(str(cell.value).strip())
        return values or None
    except Exception:
        return None


def _resolve_named_range(wb, defn):
    """Resolve a named range definition to values."""
    values = []
    try:
        for sheet_title, cell_range in defn.destinations:
            if sheet_title in wb.sheetnames:
                source = wb[sheet_title]
                for row_or_cell in source[cell_range]:
                    cells = row_or_cell if isinstance(row_or_cell, tuple) else (row_or_cell,)
                    for cell in cells:
                        if cell.value is not None and str(cell.value).strip():
                            values.append(str(cell.value).strip())
    except Exception:
        return None
    return values or None


# ---------------------------------------------------------------------------
# 7. Pipeline
# ---------------------------------------------------------------------------


@pipeline("ai-structure-proposal", timeout=3600)
@parameter(
    "excel_file",
    name="Excel File",
    type=File,
    required=True,
    help="The Excel template to analyse.",
)
@parameter(
    "gemini_connection",
    name="Gemini API Connection",
    type=CustomConnection,
    required=True,
    help="Custom connection containing the Gemini API key (field: api_key).",
)
@parameter(
    "sheet_name",
    name="Sheet Name",
    type=str,
    required=False,
    help="Specific sheet to analyse.  If empty, auto-selects the first data sheet.",
)
@parameter(
    "user_guidelines",
    name="User Guidelines",
    type=str,
    required=False,
    help=(
        "Free-text instructions to guide the analysis. "
        "E.g.: 'Rows 8-12 are a donor/funder group, not simple key-value. "
        "Columns after column Y contain province-level percentages that must "
        "sum to 100%. The primary key column is the budget line label in column B.'"
    ),
)
def ai_structure_proposal(
    excel_file: File,
    gemini_connection: CustomConnection,
    sheet_name: str = None,
    user_guidelines: str = None,
):
    """Analyse an Excel template and produce a structure_proposal.json."""
    current_run.log_info(f"Starting structure analysis of: {excel_file}")

    excel_path = excel_file.path
    api_key = gemini_connection.api_key
    model = DEFAULT_MODEL

    # -- load workbook twice: values + full (for validations) --
    current_run.log_info("Loading workbook (data_only=True for values)...")
    wb_values = openpyxl.load_workbook(str(excel_path), data_only=True)
    current_run.log_info("Loading workbook (data_only=False for validations/styles)...")
    wb_full = openpyxl.load_workbook(str(excel_path), data_only=False)

    # -- select sheet --
    ws_values, ws_full = _select_sheets(wb_values, wb_full, sheet_name)
    current_run.log_info(f"Analysing sheet: '{ws_values.title}'")

    # -- build text representation --
    current_run.log_info("Building text representation...")
    text_repr = build_text_representation(ws_values, ws_full)
    current_run.log_info(f"Text representation: {len(text_repr)} chars")

    # -- render image --
    current_run.log_info("Rendering sheet image...")
    image_bytes = render_sheet_image(ws_values)
    if image_bytes:
        current_run.log_info(f"Image rendered: {len(image_bytes)} bytes")
    else:
        current_run.log_warning("PIL not available — proceeding with text-only analysis")

    # -- deterministic validation extraction --
    current_run.log_info("Extracting data validations (deterministic)...")
    resolved_dvs = extract_resolved_validations(ws_full, wb_full)
    current_run.log_info(f"Found {len(resolved_dvs)} data validation rules")

    wb_full.close()

    # -- call Gemini --
    if user_guidelines:
        current_run.log_info(f"User guidelines: {user_guidelines[:200]}")
    prompt = build_prompt(text_repr, ws_values.title, user_guidelines)
    current_run.log_info("Sending to Gemini for analysis...")
    raw_response = call_gemini(prompt, image_bytes, api_key, model)

    # -- parse response --
    current_run.log_info("Parsing Gemini response...")
    try:
        proposal = parse_structure_json(raw_response)
    except (json.JSONDecodeError, ValueError) as exc:
        current_run.log_warning(f"JSON parse failed: {exc}")
        current_run.log_info(f"Raw response (first 2000 chars):\n{raw_response[:2000]}")
        # Save raw response for debugging
        raw_path = Path(workspace.files_path) / "structure_proposal_raw.txt"
        raw_path.write_text(raw_response, encoding="utf-8")
        current_run.add_file_output(str(raw_path))
        raise RuntimeError(
            "Gemini returned a response that could not be parsed as JSON. "
            "Raw response saved to structure_proposal_raw.txt"
        ) from exc

    # -- merge deterministic validations into proposal --
    if resolved_dvs:
        proposal["data_validations_resolved"] = resolved_dvs
        current_run.log_info(
            f"Merged {len(resolved_dvs)} deterministic validation rules into proposal"
        )

    # -- add generation metadata --
    proposal["_generation"] = {
        "generated_at": datetime.now().isoformat(),
        "model": model,
        "source_file": str(excel_path),
        "prompt_chars": len(prompt),
        "image_bytes": len(image_bytes) if image_bytes else 0,
    }

    # -- save output --
    output_path = Path(workspace.files_path) / OUTPUT_FILENAME
    current_run.log_info(f"Saving structure proposal to: {output_path}")
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(proposal, f, ensure_ascii=False, indent=2)
    current_run.add_file_output(str(output_path))

    # -- summary --
    n_sections = len(proposal.get("sections", []))
    layouts = {}
    for s in proposal.get("sections", []):
        layout = s.get("layout", "unknown")
        layouts[layout] = layouts.get(layout, 0) + 1
    current_run.log_info(
        f"Structure proposal complete: {n_sections} sections "
        f"({', '.join(f'{v} {k}' for k, v in layouts.items())})"
    )

    wb_values.close()


def _select_sheets(wb_values, wb_full, sheet_name=None):
    """Select the worksheet to analyse from both workbooks."""
    skip = {"guide", "liste des catégories", "liste des categories", "observations"}

    if sheet_name:
        return wb_values[sheet_name], wb_full[sheet_name]

    for name in wb_values.sheetnames:
        if name.strip().lower() not in skip:
            return wb_values[name], wb_full[name]

    return wb_values[wb_values.sheetnames[0]], wb_full[wb_full.sheetnames[0]]


if __name__ == "__main__":
    ai_structure_proposal()
