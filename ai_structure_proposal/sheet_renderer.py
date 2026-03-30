"""Sheet-to-PNG renderer for multimodal LLM analysis.

Renders the top-left region of an Excel worksheet as a PNG image using
Pillow.  The image gives the LLM a visual gestalt of the layout — colour
bands, merged-header groups, density shifts — that complements the text
representation.
"""

import io

from openpyxl.utils import get_column_letter

from excel_reader import ExcelReader
from PIL import ImageFont, Image, ImageDraw

# Default rendering parameters.
MAX_ANALYSIS_ROWS = 35
MAX_ANALYSIS_COLS = 60


class SheetRenderer:
    """Renders a worksheet region to a PNG image using Pillow.

    The renderer is stateless — it can be reused across multiple sheets.

    Args:
        max_rows: Maximum number of rows to include in the image.
        max_cols: Maximum number of columns to include in the image.
    """

    # Cell dimensions in pixels.
    CELL_WIDTH = 90
    ROW_HEIGHT = 20
    HEADER_HEIGHT = 22
    LABEL_WIDTH = 35

    def __init__(
        self,
        max_rows: int = MAX_ANALYSIS_ROWS,
        max_cols: int = MAX_ANALYSIS_COLS,
    ):
        """Initialise the renderer with row/column limits.

        Args:
            max_rows: How many rows (from the top) to render.
            max_cols: How many columns (from the left) to render.
        """
        self.max_rows = max_rows
        self.max_cols = max_cols

    def render(self, worksheet) -> bytes | None:
        """Render the top-left region of *worksheet* as a PNG.

        Returns:
            The PNG image as ``bytes``, or ``None`` if Pillow is not installed.
        """

        actual_cols = min(worksheet.max_column or 1, self.max_cols)
        actual_rows = min(worksheet.max_row or 1, self.max_rows)

        image_width = self.LABEL_WIDTH + actual_cols * self.CELL_WIDTH + 2
        image_height = self.HEADER_HEIGHT + actual_rows * self.ROW_HEIGHT + 2

        image = Image.new("RGB", (image_width, image_height), "#FFFFFF")
        draw = ImageDraw.Draw(image)

        font_regular, font_bold = self._load_fonts()

        # -- merge helpers --
        skip_cells: set[tuple[int, int]] = set()
        merge_spans: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in worksheet.merged_cells.ranges:
            merge_spans[(merged_range.min_row, merged_range.min_col)] = (
                merged_range.max_row - merged_range.min_row + 1,
                merged_range.max_col - merged_range.min_col + 1,
            )
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != (merged_range.min_row, merged_range.min_col):
                        skip_cells.add((row, col))

        # -- column headers --
        for col in range(1, actual_cols + 1):
            x_pos = self.LABEL_WIDTH + (col - 1) * self.CELL_WIDTH
            draw.rectangle(
                [x_pos, 0, x_pos + self.CELL_WIDTH, self.HEADER_HEIGHT],
                fill="#E8E8E8",
                outline="#CCCCCC",
            )
            draw.text(
                (x_pos + 4, 4),
                get_column_letter(col),
                fill="#333333",
                font=font_bold,
            )

        # -- cell grid --
        for row in range(1, actual_rows + 1):
            y_pos = self.HEADER_HEIGHT + (row - 1) * self.ROW_HEIGHT

            # row label
            draw.rectangle(
                [0, y_pos, self.LABEL_WIDTH, y_pos + self.ROW_HEIGHT],
                fill="#E8E8E8",
                outline="#CCCCCC",
            )
            draw.text((3, y_pos + 3), str(row), fill="#555555", font=font_regular)

            for col in range(1, actual_cols + 1):
                if (row, col) in skip_cells:
                    continue

                x_pos = self.LABEL_WIDTH + (col - 1) * self.CELL_WIDTH

                # cell dimensions (handle merged cells)
                if (row, col) in merge_spans:
                    row_span, col_span = merge_spans[(row, col)]
                    cell_width = min(col_span, actual_cols - col + 1) * self.CELL_WIDTH
                    cell_height = min(row_span, actual_rows - row + 1) * self.ROW_HEIGHT
                else:
                    cell_width = self.CELL_WIDTH
                    cell_height = self.ROW_HEIGHT

                cell = worksheet.cell(row, col)
                background_color = ExcelReader._cell_background_hex(cell) or "#FFFFFF"
                text_color = "#000000"

                # auto-contrast for dark backgrounds
                try:
                    red = int(background_color[1:3], 16)
                    green = int(background_color[3:5], 16)
                    blue = int(background_color[5:7], 16)
                    if (red + green + blue) / 3 < 128:
                        text_color = "#FFFFFF"
                except (ValueError, IndexError):
                    pass

                draw.rectangle(
                    [x_pos, y_pos, x_pos + cell_width, y_pos + cell_height],
                    fill=background_color,
                    outline="#CCCCCC",
                )

                if cell.value is not None:
                    text = str(cell.value)
                    max_chars = max((cell_width - 8) // 6, 4)
                    if len(text) > max_chars:
                        text = text[: max_chars - 2] + ".."
                    chosen_font = font_bold if (cell.font and cell.font.bold) else font_regular
                    draw.text(
                        (x_pos + 4, y_pos + 3),
                        text,
                        fill=text_color,
                        font=chosen_font,
                    )

        # -- encode as PNG bytes --
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()

    @staticmethod
    def _load_fonts():
        """Load DejaVu Sans fonts, falling back to Pillow's built-in default.

        Returns:
            A tuple ``(font_regular, font_bold)``.
        """
        font_paths = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/TTF/DejaVuSans.ttf",
        ]
        for path in font_paths:
            try:
                return (
                    ImageFont.truetype(path, 10),
                    ImageFont.truetype(path.replace("Sans.", "Sans-Bold."), 10),
                )
            except OSError:
                continue

        default_font = ImageFont.load_default()
        return default_font, default_font
