"""OpenHEXA Pipeline: Generate schema_validation.json from structure_proposal.json.

Reads a structure_proposal.json (produced by the AI structure analysis pipeline)
and an Excel template, then generates a schema_validation.json that the
validation pipeline can consume.

This pipeline is fully deterministic — it reshapes the AI-proposed structure
into the schema format, enriches it with example values and resolved data
validations from the actual Excel file.

No hardcoded layout heuristics: all structural decisions come from the
structure_proposal.json, which can be reviewed and edited by the user.
"""

import datetime
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from openhexa.sdk import File, current_run, parameter, pipeline, workspace
from openpyxl.utils import column_index_from_string, get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCHEMA_OUTPUT_FILENAME = "schema_validation.json"


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------


_LIGATURES = str.maketrans({"œ": "oe", "Œ": "OE", "æ": "ae", "Æ": "AE"})


def to_canonical(text):
    """Convert header text to a snake_case canonical name."""
    if not text:
        return "unnamed"
    text = str(text).strip()
    nfkd = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in nfkd if not unicodedata.combining(c))
    text = text.translate(_LIGATURES)
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_").lower()
    return text or "unnamed"


def to_column_canonical(text):
    """Deterministic canonical name for a column header.

    Strips parenthetical descriptions before canonicalizing so that
    "Niveau central (Directions, Programmes, …)" → "niveau_central"
    rather than an unwieldy 99-character slug.
    """
    text = re.sub(r"\s*\([^)]*\)", "", str(text or "")).strip()
    return to_canonical(text)


def col_idx(letter):
    """Column letter → 1-based index."""
    return column_index_from_string(letter)


def col_letter(idx):
    """1-based index → column letter."""
    return get_column_letter(idx)


def cell_ref_to_col(cell_ref):
    """Extract column letter from a cell reference like 'C7' or 'AB12'."""
    return "".join(c for c in cell_ref if c.isalpha())


def cell_ref_to_row(cell_ref):
    """Extract row number from a cell reference like 'C7'."""
    return int("".join(c for c in cell_ref if c.isdigit()))


def format_example(value):
    """Format a cell value as an example string."""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def infer_type(value):
    """Infer schema type from a Python value."""
    if value is None:
        return "string"
    if isinstance(value, datetime.datetime):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    s = str(value).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return "date"
    try:
        float(s.replace(",", "").replace(" ", ""))
        return "number"
    except (ValueError, AttributeError):
        pass
    return "string"


# ---------------------------------------------------------------------------
# Data Validation extraction (from Excel, deterministic)
# ---------------------------------------------------------------------------


def extract_data_validations(ws):
    """Extract data validation rules → dict mapping cell ref → dv info.

    Also captures INDIRECT formulas that couldn't be resolved, storing them
    with type='indirect' so they can be enriched later by
    resolve_indirect_dependencies().
    """
    dv_map = {}
    if not hasattr(ws, "data_validations") or ws.data_validations is None:
        return dv_map

    wb = ws.parent
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue

        formula = dv.formula1
        if not formula:
            continue
        formula_str = str(formula).strip('"')

        # Check if this is an INDIRECT formula
        is_indirect = formula_str.upper().startswith("INDIRECT(")
        choices = _resolve_dv_formula(ws, wb, formula_str)

        if is_indirect and not choices:
            # Store as unresolved INDIRECT — will be enriched later
            parent_col = _parse_indirect_parent_col(formula_str)
            if parent_col:
                dv_info = {
                    "type": "indirect",
                    "depends_on_column": parent_col,
                    "formula": formula_str,
                    "allow_blank": bool(dv.allow_blank),
                }
            else:
                continue
        elif choices:
            dv_info = {
                "type": "list",
                "values": choices,
                "allow_blank": bool(dv.allow_blank),
                "formula": formula_str,
            }
        else:
            continue

        for cell_range in dv.sqref.ranges:
            for row in range(cell_range.min_row, cell_range.max_row + 1):
                for col in range(cell_range.min_col, cell_range.max_col + 1):
                    ref = f"{get_column_letter(col)}{row}"
                    dv_map[ref] = dv_info

    return dv_map


def _parse_indirect_parent_col(formula_str):
    """Extract the parent column letter from an INDIRECT formula.

    E.g. INDIRECT($L17) → 'L', INDIRECT($N17) → 'N'.
    """
    inner = re.search(r"INDIRECT\(\$?([A-Z]+)\d+\)", formula_str, re.IGNORECASE)
    if inner:
        return inner.group(1).upper()
    return None


def resolve_indirect_dependencies(ws, dv_map):
    """Resolve INDIRECT dependencies into conditional_enum entries.

    For each 'indirect' entry in dv_map, finds the parent column's allowed
    values and resolves each as a named range to build a values_by_parent map.

    Returns a dict mapping child column letter → conditional validation info.
    """
    wb = ws.parent
    conditional_map = {}

    # Collect all indirect entries grouped by child column
    indirect_cols = {}
    for ref, info in dv_map.items():
        if info.get("type") != "indirect":
            continue
        child_col = "".join(c for c in ref if c.isalpha())
        if child_col not in indirect_cols:
            indirect_cols[child_col] = info

    for child_col, info in indirect_cols.items():
        parent_col = info["depends_on_column"]

        # Find parent column's allowed values from dv_map
        parent_values = _find_column_values(dv_map, parent_col)
        if not parent_values:
            continue

        # For each parent value, try to resolve it as a named range
        # (with common suffix variants: value, value_, value..)
        values_by_parent = {}
        for pval in parent_values:
            resolved = _resolve_named_range_variants(wb, pval)
            if resolved:
                values_by_parent[pval] = resolved

        if values_by_parent:
            conditional_map[child_col] = {
                "depends_on_column": parent_col,
                "values_by_parent": values_by_parent,
            }

    return conditional_map


def _find_column_values(dv_map, col_letter):
    """Find the allowed values for a column from the dv_map.

    Looks for any cell in that column with a 'list' type validation.
    """
    for ref, info in dv_map.items():
        ref_col = "".join(c for c in ref if c.isalpha())
        if ref_col == col_letter and info.get("type") == "list":
            return info.get("values", [])
    return []


def _resolve_named_range_variants(wb, name):
    """Try to resolve a named range with common suffix variants.

    Excel templates often use suffixes like '_' or '.' when the parent
    value contains spaces or special characters.
    """
    if not hasattr(wb, "defined_names"):
        return []

    # Try exact name first
    result = _resolve_named_range(wb, name)
    if result:
        return result

    # Try common variants: with underscore suffix, stripped spaces, etc.
    variants = [
        name + "_",
        name.replace(" ", "_"),
        name.replace(" ", "_") + "_",
        name.replace("'", ""),
        name.replace("'", "") + "_",
        re.sub(r"[^a-zA-Z0-9_àâäéèêëïîôùûüÿçœæ]", "_", name),
        re.sub(r"[^a-zA-Z0-9_àâäéèêëïîôùûüÿçœæ]", "_", name) + "_",
    ]
    # Deduplicate while preserving order
    seen = {name}
    for v in variants:
        if v not in seen:
            seen.add(v)
            result = _resolve_named_range(wb, v)
            if result:
                return result

    return []


def _resolve_dv_formula(ws, wb, formula_str):
    """Resolve a data validation formula to allowed values list."""
    if formula_str.upper().startswith("INDIRECT("):
        return _resolve_indirect(ws, wb, formula_str)

    if "$" in formula_str or ":" in formula_str or "!" in formula_str:
        return _resolve_range_values(ws, wb, formula_str)

    if formula_str.isidentifier() or (
        formula_str.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .isidentifier()
    ):
        resolved = _resolve_named_range(wb, formula_str)
        if resolved:
            return resolved

    if ";" in formula_str:
        return [v.strip() for v in formula_str.split(";") if v.strip()]
    if "," in formula_str:
        return [v.strip() for v in formula_str.split(",") if v.strip()]

    return [formula_str.strip()] if formula_str.strip() else []


def _resolve_named_range(wb, name):
    """Resolve a named range to its values."""
    if not hasattr(wb, "defined_names"):
        return []
    defn = wb.defined_names.get(name)
    if defn is None:
        return []
    values = []
    try:
        for sheet_title, cell_range in defn.destinations:
            if sheet_title in wb.sheetnames:
                source_ws = wb[sheet_title]
                for row_or_cell in source_ws[cell_range]:
                    row_cells = row_or_cell if isinstance(row_or_cell, tuple) else (row_or_cell,)
                    for cell in row_cells:
                        if cell.value is not None and str(cell.value).strip():
                            values.append(str(cell.value).strip())
    except Exception:
        return []
    return values


def _resolve_indirect(ws, wb, formula_str):
    """Best-effort INDIRECT() resolution."""
    inner = re.search(r"INDIRECT\((.+)\)", formula_str, re.IGNORECASE)
    if not inner:
        return []
    ref = inner.group(1).replace("$", "").strip("'\"")
    try:
        cell_val = ws[ref].value
        if cell_val and isinstance(cell_val, str):
            return _resolve_named_range(wb, cell_val.strip())
    except Exception:
        pass
    return []


def _resolve_range_values(ws, wb, formula_str):
    """Resolve a cell range reference to values."""
    values = []
    try:
        if "!" in formula_str:
            sheet_part, range_part = formula_str.rsplit("!", 1)
            sheet_name = sheet_part.strip("'\"")
            source_ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if source_ws is None:
                return values
        else:
            source_ws = ws
            range_part = formula_str
        range_part = range_part.replace("$", "")
        for row_or_cell in source_ws[range_part]:
            row_cells = row_or_cell if isinstance(row_or_cell, tuple) else (row_or_cell,)
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    values.append(str(cell.value).strip())
    except Exception:
        pass
    return values


def _merge_proposal_validations(dv_map, proposal):
    """Enrich dv_map with pre-resolved values from the proposal's data_validations_resolved.

    Only adds entries that are missing from dv_map (i.e., the Excel-based
    resolution failed or couldn't resolve INDIRECT/named-range formulas).
    """
    resolved_list = proposal.get("data_validations_resolved", [])
    for entry in resolved_list:
        vals = entry.get("resolved_values")
        if not vals:
            continue
        sqref = entry.get("cells", "")
        dv_info = {
            "type": entry.get("type", "list"),
            "values": vals,
            "allow_blank": entry.get("allow_blank", True),
            "formula": entry.get("formula", ""),
        }
        # Parse sqref like "C7" or "J17:J334"
        for cell_ref in _expand_sqref(sqref):
            if cell_ref not in dv_map:
                dv_map[cell_ref] = dv_info


def _expand_sqref(sqref):
    """Expand a sqref like 'J17:J334' into individual cell refs.

    For efficiency, only expands the first 500 cells of a range.
    """
    if ":" not in sqref:
        return [sqref.strip()]

    parts = sqref.strip().split(":")
    if len(parts) != 2:
        return [sqref.strip()]

    start, end = parts
    start_col = "".join(c for c in start if c.isalpha())
    start_row = int("".join(c for c in start if c.isdigit()) or "0")
    end_col = "".join(c for c in end if c.isalpha())
    end_row = int("".join(c for c in end if c.isdigit()) or "0")

    if not start_row or not end_row:
        return [sqref.strip()]

    refs = []
    s_idx = column_index_from_string(start_col)
    e_idx = column_index_from_string(end_col)
    for ci in range(s_idx, e_idx + 1):
        cl = get_column_letter(ci)
        for ri in range(start_row, min(end_row + 1, start_row + 500)):
            refs.append(f"{cl}{ri}")
    return refs


# ---------------------------------------------------------------------------
# Section processors: structure_proposal → schema parts
# ---------------------------------------------------------------------------


def _get_module_key_from_section(section):
    """Derive a module_key like 'module_1' from the section title if possible."""
    title = (section.get("title") or {}).get("expected", "")
    m = re.search(r"module\s*(\d+)", title, re.IGNORECASE)
    if m:
        return f"module_{m.group(1)}"
    return section.get("key", "unknown")


def process_key_value(section, ws, dv_map):
    """Process a key_value section → (labels_dict, values_dict)."""
    mod_key = _get_module_key_from_section(section)

    labels = {}
    values = {}

    # Title label
    title_info = section.get("title") or {}
    if title_info.get("cell"):
        labels["title"] = {
            "cell": title_info["cell"],
            "expected": title_info.get("expected", ""),
        }

    for field in section.get("fields", []):
        fkey = field.get("key", "unnamed")

        # Label
        label_info = field.get("label", {})
        label_cell = label_info.get("cell", "")
        labels[fkey] = {
            "cell": label_cell,
            "expected": label_info.get("expected", ""),
        }

        # Value
        value_info = field.get("value", {})
        value_cell = value_info.get("cell", "")
        # Extract the first cell from ranges like "C3:H3" → "C3"
        value_cell_single = value_cell.split(":")[0] if ":" in value_cell else value_cell

        value_type = value_info.get("type", "string")

        # Read example from Excel
        example = ""
        if value_cell_single:
            try:
                col_l = cell_ref_to_col(value_cell_single)
                row_n = cell_ref_to_row(value_cell_single)
                raw_val = ws.cell(row=row_n, column=col_idx(col_l)).value
                example = format_example(raw_val)
            except Exception:
                pass

        label_display = label_info.get("expected", "").rstrip(":").strip()
        value_key = f"{fkey}_value"
        value_entry = {
            "cell": value_cell_single,
            "label_ref": f"{mod_key}.{fkey}",
            "label_display": label_display,
            "type": value_type,
            "example": example,
            "required": False,
            "message": (f"Le champ '{label_display}' (cellule {value_cell_single}) est optionnel."),
        }

        # Attach enum from dv_map
        if value_cell_single and dv_map.get(value_cell_single):
            dv_info = dv_map[value_cell_single]
            if dv_info.get("values"):
                value_entry["validation"] = {"enum": dv_info["values"]}
                value_entry["type"] = "string"

        values[value_key] = value_entry

    return mod_key, labels, values


def process_grouped_key_value(section, ws, dv_map):
    """Process a grouped_key_value section → (labels_dict, values_dict, element_columns).

    Returns (section_key, mod_key, labels, values, column_labels_entry).
    """
    section_key = section.get("key", "group")
    mod_key = _get_module_key_from_section(section)

    labels = {}
    values = {}

    # Title label
    title_info = section.get("title") or {}
    section_title = title_info.get("expected", section_key)
    if title_info.get("cell"):
        fkey = to_canonical(section_title)
        labels[fkey] = {
            "cell": title_info["cell"],
            "expected": section_title,
        }

    # Elements (the column headers — e.g. numbered entities)
    elements = section.get("elements", {})
    element_cols = elements.get("columns", [])
    element_labels = elements.get("labels", [])
    header_row = elements.get("header_row")

    # Build element column labels entry
    column_labels_items = []
    for i, col_l in enumerate(element_cols):
        label_text = element_labels[i] if i < len(element_labels) else f"Element {i + 1}"
        column_labels_items.append(
            {
                "cell": f"{col_l}{header_row}",
                "expected": label_text,
            }
        )

    column_labels_entry = None
    if column_labels_items:
        display_name = (
            section_title
            if section_title and section_title != section_key
            else section_key.replace("_", " ").title()
        )
        column_labels_entry = {
            "_description": f"Column labels for grouped structure '{section_key}' (row {header_row}).",
            "layout": "grouped_key_value",
            "display_name": display_name,
            "count_is_dynamic": elements.get("count_is_dynamic", False),
            "header_row": header_row,
            "items": column_labels_items,
        }

    # Property rows (one per field in the grouped structure)
    property_rows = {}
    for i, field in enumerate(section.get("fields", [])):
        fkey = field.get("key", "unnamed")

        # Label entry
        label_info = field.get("label", {})
        row_n = label_info.get("row")
        col_l = label_info.get("col", "B")
        label_cell = f"{col_l}{row_n}" if row_n else ""
        labels[fkey] = {
            "cell": label_cell,
            "expected": label_info.get("expected", ""),
        }

        # Read example from first element column
        example = ""
        first_col = element_cols[0] if element_cols else None
        if first_col and row_n:
            try:
                raw_val = ws.cell(row=row_n, column=col_idx(first_col)).value
                example = format_example(raw_val)
            except Exception:
                pass

        is_first = i == 0
        field_type = field.get("type", "string")
        label_display = label_info.get("expected", "")

        row_entry = {
            "row": row_n,
            "label_ref": f"{mod_key}.{fkey}",
            "label_display": label_display,
            "type": field_type,
            "example": example,
            "required_rule": field.get(
                "required_rule",
                "anchor" if is_first else "if_active",
            ),
            "message": (
                f"Au moins un '{label_display}' (ligne {row_n}) doit être renseigné."
                if is_first
                else f"Le '{label_display}' (ligne {row_n}) est requis pour chaque élément actif."
            ),
        }

        property_rows[fkey] = row_entry

    # Pack grouped values into values dict using the section key
    grouped_key = f"{section_key}_values"
    values[grouped_key] = {
        "_description": f"Grouped rows for '{section_key}' under columns {', '.join(element_cols)}.",
        "columns": element_cols,
        "rows": property_rows,
    }

    return section_key, mod_key, labels, values, column_labels_entry


def process_records(section, ws, dv_map, data_start_row, conditional_map=None):
    """Process a records section → header_module dict.

    Returns (mod_key, header_module_entry, title_label_entry).
    """
    conditional_map = conditional_map or {}
    mod_key = _get_module_key_from_section(section)
    title_info = section.get("title") or {}
    headers_info = section.get("headers") or {}

    result = {
        "title": {
            "cell": title_info.get("cell", ""),
            "expected": title_info.get("expected", ""),
        },
        "start_column": section.get("area", {}).get("start_col", ""),
        "columns": [],
    }

    # Process columns
    for col_def in headers_info.get("columns", []):
        col_type = col_def.get("type", col_def.get("position", "single"))

        if col_type in ("single", "fixed"):
            result["columns"].append(
                _build_fixed_column(col_def, ws, dv_map, data_start_row, conditional_map)
            )

        elif col_type == "group":
            result["columns"].append(_build_group_column(col_def, ws, dv_map, data_start_row))

    # Title label for module_headers_rowN
    title_label = {
        "cell": title_info.get("cell", ""),
        "expected": title_info.get("expected", ""),
    }

    return mod_key, result, title_label


def _build_fixed_column(col_def, ws, dv_map, data_start_row, conditional_map=None):
    """Build a fixed (single) column entry for header_modules."""
    conditional_map = conditional_map or {}
    col_l = col_def.get("column", "")
    header_row = col_def.get("header_row", 15)

    # Read example value from first data row
    example = ""
    if col_l and data_start_row:
        try:
            raw_val = ws.cell(row=data_start_row, column=col_idx(col_l)).value
            example = format_example(raw_val)
        except Exception:
            pass

    entry = {
        "position": "fixed",
        "column": col_l,
        "row": header_row,
        "expected": col_def.get("expected", ""),
        "canonical_name": to_column_canonical(col_def.get("expected", "")),
        "value_type": col_def.get("value_type", "string"),
        "example": example,
        "required": col_def.get("required", False),
        "known_variants": [],
    }

    # Attach validation from dv_map (check data cell in this column)
    data_cell_ref = f"{col_l}{data_start_row}"

    # Check if this column has a conditional dependency (INDIRECT dropdown)
    if col_l in conditional_map:
        cond = conditional_map[col_l]
        entry["validation"] = {
            "conditional_enum": {
                "depends_on_column": cond["depends_on_column"],
                "values_by_parent": cond.get("values_by_parent", {}),
            },
        }
        entry["value_type"] = "string"
    elif dv_map.get(data_cell_ref):
        dv_info = dv_map[data_cell_ref]
        if dv_info.get("values"):
            entry["validation"] = {"enum": dv_info["values"]}
            entry["value_type"] = "string"

    return entry


def _build_group_column(col_def, ws, dv_map, data_start_row):
    """Build a group column entry for header_modules."""
    parent = col_def.get("parent", {})
    group = {
        "position": "group",
        "parent": {
            "row": parent.get("header_row", 15),
            "expected": parent.get("expected", ""),
            "known_variants": [],
        },
        "sub_columns": [],
    }

    parent_canonical = to_column_canonical(parent.get("expected", ""))

    for sub in col_def.get("sub_columns", []):
        # Handle both new format (fixed: true/false) and old format (position: fixed/dynamic)
        is_fixed = sub.get("fixed", sub.get("position") == "fixed")
        is_dynamic = not is_fixed if "fixed" in sub else sub.get("position") == "dynamic"

        if is_dynamic:
            # Dynamic sub-column: always use parent canonical as prefix
            group["sub_columns"].append(
                {
                    "position": "dynamic",
                    "row": sub.get("header_row", 16),
                    "pattern": sub.get("pattern", ".*"),
                    "canonical_prefix": parent_canonical,
                    "value_type": sub.get("value_type", "string"),
                    "description": sub.get("description", ""),
                }
            )
        else:
            # Fixed sub-column: always {parent}_{to_column_canonical(header)}
            sub_col = sub.get("column", "")
            sub_canonical = f"{parent_canonical}_{to_column_canonical(sub.get('expected', ''))}"

            example = ""
            if sub_col and data_start_row:
                try:
                    raw_val = ws.cell(row=data_start_row, column=col_idx(sub_col)).value
                    example = format_example(raw_val)
                except Exception:
                    pass

            sub_entry = {
                "position": "fixed",
                "column": sub_col,
                "row": sub.get("header_row", 16),
                "expected": sub.get("expected", ""),
                "canonical_name": sub_canonical,
                "value_type": sub.get("value_type", "string"),
                "example": example,
                "required": sub.get("required", False),
                "known_variants": [],
            }

            # Attach validation from dv_map
            data_cell_ref = f"{sub_col}{data_start_row}"
            if sub_col and dv_map.get(data_cell_ref):
                dv_info = dv_map[data_cell_ref]
                if dv_info.get("values"):
                    sub_entry["validation"] = {"enum": dv_info["values"]}
                    sub_entry["value_type"] = "string"

            group["sub_columns"].append(sub_entry)

    return group


def _find_primary_key(hdr_entry: dict, headers_proposal: dict) -> str | None:
    """Resolve the primary_key from processed header columns.

    Looks up Gemini's primary_key suggestion in the deterministic column list.
    Falls back to the first required column, then the first column overall.
    """
    gemini_pk = headers_proposal.get("primary_key", "")
    columns = hdr_entry.get("columns", [])

    # Flatten all fixed columns (including inside groups)
    flat: list[dict] = []
    for col in columns:
        if col.get("position") == "fixed":
            flat.append(col)
        elif col.get("position") == "group":
            for sub in col.get("sub_columns", []):
                if sub.get("position") == "fixed":
                    flat.append(sub)

    # Try to match Gemini's suggestion to a deterministic canonical
    if gemini_pk:
        for col in flat:
            expected = col.get("expected", "")
            if to_column_canonical(expected) == to_column_canonical(gemini_pk):
                return col.get("canonical_name")
            if col.get("canonical_name") == gemini_pk:
                return col["canonical_name"]

    # Fallback: first required column
    for col in flat:
        if col.get("required"):
            return col.get("canonical_name")

    # Last resort: first column
    if flat:
        return flat[0].get("canonical_name")

    return None


# ---------------------------------------------------------------------------
# Table definitions builder
# ---------------------------------------------------------------------------


def build_table_definitions(
    label_fields, value_fields, meta_module_keys, header_row, data_start_row, primary_key=None
):
    """Build the table_definitions section."""
    td = {}

    # Metadata table: collect all label → value field pairs from metadata modules
    fields_list = []
    for mod_key in meta_module_keys:
        labels = label_fields.get(mod_key, {})
        vals = value_fields.get(f"{mod_key}_values", {})

        for vk, vf in vals.items():
            if vk.startswith("_"):
                continue
            # Skip grouped_key_value entries (they end with _values
            # and contain "columns"/"rows" instead of "cell")
            if isinstance(vf, dict) and "columns" in vf and "rows" in vf:
                continue
            if not isinstance(vf, dict) or "cell" not in vf:
                continue
            # Derive label_key from label_ref
            label_key = (
                vf.get("label_ref", "").split(".")[-1] if "." in vf.get("label_ref", "") else ""
            )
            label_cell = labels.get(label_key, {}).get("cell", "")
            fields_list.append(
                {
                    "canonical_name": label_key or to_canonical(vf.get("label_display", vk)),
                    "label_cell": label_cell,
                    "value_cell": vf.get("cell", ""),
                    "type": vf.get("type", "string"),
                }
            )

    # Grouped fields from metadata modules (e.g., repeating column groups)
    grouped_fields = {}
    for mod_key in meta_module_keys:
        vals = value_fields.get(f"{mod_key}_values", {})
        for vk, vf in vals.items():
            if not isinstance(vf, dict):
                continue
            if "columns" not in vf or "rows" not in vf:
                continue
            # This is a grouped_key_value entry
            sec_key = vk.removesuffix("_values") if vk.endswith("_values") else vk
            grouped_fields[f"{sec_key}_fields"] = {
                "_description": (
                    f"Repeating group '{sec_key}' (columns {', '.join(vf['columns'])})."
                ),
                "columns": vf["columns"],
                "fields_per_column": [
                    {
                        "canonical_name": f"{sec_key}_{rk}",
                        "row": rd["row"],
                        "type": rd["type"],
                    }
                    for rk, rd in vf.get("rows", {}).items()
                ],
            }

    td["metadata"] = {
        "_description": "Metadata table built from label-value pairs.",
        "source": "metadata sections",
        "fields": fields_list,
    }
    td["metadata"].update(grouped_fields)

    # Data table
    td["records_data"] = {
        "_description": "Data table built from data rows. Each row becomes a record.",
        "header_row_start": header_row,
        "header_row_end": header_row + 1,
        "data_row_start": data_start_row,
        "data_row_end": "dynamic",
        "header_source": "header_modules",
        "empty_row_terminates": True,
        "row_validation": {
            (primary_key or "primary_column"): {
                "required": True,
                "message": (
                    f"Chaque ligne de données doit avoir un '{primary_key or 'primary_column'}'."
                ),
            },
        },
    }

    return td


# ---------------------------------------------------------------------------
# Validation config builder
# ---------------------------------------------------------------------------


def build_validation_config(data_start_row):
    """Build the validation_config section."""
    return {
        "_description": "Configuration for validation and header matching behavior.",
        "label_comparison": {
            "case_sensitive": False,
            "strip_whitespace": True,
            "strip_trailing_colon": True,
            "normalize_accents": True,
        },
        "header_matching": {
            "strategy": "exact_then_variants_then_ask",
            "on_mismatch": "warn_and_ask_confirmation",
            "on_missing": "error",
        },
        "dynamic_columns": {
            "_description": (
                "Columns with position='dynamic' are matched by regex pattern. "
                "Canonical names are generated as '{canonical_prefix}_{normalized_value}'."
            ),
        },
        "data_rows": {
            "start_row": data_start_row,
            "termination": "first_fully_empty_row",
        },
    }


# ---------------------------------------------------------------------------
# Main schema generator
# ---------------------------------------------------------------------------


def generate_schema(proposal, ws, dv_map, conditional_map=None):
    """Transform a structure_proposal + Excel data into schema_validation."""
    conditional_map = conditional_map or {}
    landmarks = proposal.get("structure_landmarks", {})
    title_row = landmarks.get("title_row", 14)
    header_rows = landmarks.get("header_rows", [15, 16])
    data_start_row = landmarks.get("data_start_row", 17)
    header_row = header_rows[0] if header_rows else 15

    label_fields = {}
    value_fields = {}
    header_modules = {
        "_description": (
            "Each module defines its own header columns independently. "
            "Columns can be 'fixed', 'dynamic', or 'group'."
        ),
        "mismatch_message_template": (
            "L'en-tête en colonne {column} ('{actual_value}') ne correspond pas "
            "à la valeur attendue ('{expected}'). Confirmez-vous que cette colonne "
            "correspond bien à '{canonical_name}' ?"
        ),
    }
    module_headers_row = {}
    meta_module_keys = []
    primary_key = None  # Will be read from the first records section

    for section in proposal.get("sections", []):
        layout = section.get("layout")

        if layout == "key_value":
            mod_key, labels, values = process_key_value(section, ws, dv_map)
            label_fields[mod_key] = labels
            value_fields[f"{mod_key}_values"] = values
            meta_module_keys.append(mod_key)
            current_run.log_info(
                f"  key_value section '{mod_key}': {len(labels)} labels, {len(values)} values"
            )

        elif layout == "grouped_key_value":
            sec_key, mod_key, labels, values, col_labels = process_grouped_key_value(
                section, ws, dv_map
            )
            # Merge labels into the parent module's label_fields
            parent_key = _find_parent_module(meta_module_keys) or mod_key
            if parent_key in label_fields:
                label_fields[parent_key].update(labels)
            else:
                label_fields[mod_key] = labels
                meta_module_keys.append(mod_key)

            # Rewrite label_ref in grouped rows to use parent key
            grouped_key = f"{sec_key}_values"
            if parent_key != mod_key:
                gv = values.get(grouped_key, {})
                for _, rd in gv.get("rows", {}).items():
                    if "label_ref" in rd:
                        rd["label_ref"] = rd["label_ref"].replace(
                            f"{mod_key}.", f"{parent_key}.", 1
                        )

            # Merge values into parent values
            parent_val_key = f"{parent_key}_values"
            if parent_val_key in value_fields:
                value_fields[parent_val_key].update(values)
            else:
                value_fields[f"{mod_key}_values"] = values
                if mod_key not in meta_module_keys:
                    meta_module_keys.append(mod_key)

            if col_labels:
                # Add references to the values section so the UI can find linked data
                actual_parent = parent_key if parent_key in label_fields else mod_key
                col_labels["values_section"] = f"{actual_parent}_values"
                col_labels["values_key"] = grouped_key
                col_labels["label_module"] = actual_parent
                label_fields[f"{sec_key}_columns"] = col_labels

            current_run.log_info(
                f"  grouped_key_value section '{sec_key}': "
                f"{len(labels)} labels, merged into '{parent_key}'"
            )

        elif layout == "records":
            mod_key, hdr_entry, title_label = process_records(
                section, ws, dv_map, data_start_row, conditional_map
            )
            header_modules[mod_key] = hdr_entry
            module_headers_row[mod_key] = title_label

            # Capture primary_key from the first records section
            # Use the deterministic canonical_name from the processed columns
            # (never trust Gemini's raw primary_key value).
            if primary_key is None:
                primary_key = _find_primary_key(
                    hdr_entry,
                    section.get("headers", {}),
                )

            n_cols = len(hdr_entry.get("columns", []))
            current_run.log_info(f"  records section '{mod_key}': {n_cols} column entries")

    # Add module_headers_rowN
    if module_headers_row:
        row_key = f"module_headers_row{title_row}"
        label_fields[row_key] = module_headers_row

    # Deduplicate canonical names that collide across modules (e.g. "NA")
    _deduplicate_canonicals(header_modules)

    # Build table_definitions
    table_defs = build_table_definitions(
        label_fields,
        value_fields,
        meta_module_keys,
        header_row,
        data_start_row,
        primary_key,
    )

    # Assemble schema
    schema = {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "version": "1.0.0",
        "description": (
            f"Validation schema for '{proposal.get('source_sheet', '')}' sheet "
            f"— generated from structure_proposal.json."
        ),
        "sheet_name": proposal.get("source_sheet", ""),
        "label_fields": label_fields,
        "value_fields": value_fields,
        "header_modules": header_modules,
        "table_definitions": table_defs,
        "validation_config": build_validation_config(data_start_row),
    }

    # Carry over template_file if present
    gen = proposal.get("_generation", {})
    if gen.get("source_file"):
        schema["template_file"] = gen["source_file"]

    return schema


def _deduplicate_canonicals(header_modules):
    """Detect canonical name collisions across modules and prefix with module key.

    Example: Modules 6, 7, and 8 all have an "NA" column → to_column_canonical
    produces "na" for all three.  After deduplication they become "module_6_na",
    "module_7_na", "module_8_na".
    """
    name_map: dict[str, list[tuple[str, dict]]] = {}
    for mod_key, mod_def in header_modules.items():
        if mod_key.startswith("_") or mod_key == "mismatch_message_template":
            continue
        if not isinstance(mod_def, dict):
            continue
        for col_entry in mod_def.get("columns", []):
            _collect_canonicals(name_map, mod_key, col_entry)

    for name, entries in name_map.items():
        if len(entries) > 1:
            for mod_key, entry_dict in entries:
                entry_dict["canonical_name"] = f"{mod_key}_{name}"


def _collect_canonicals(name_map, mod_key, col_entry):
    """Walk a column entry and collect canonical_name → (mod_key, entry) pairs."""
    pos = col_entry.get("position", "")
    if pos == "fixed":
        canonical = col_entry.get("canonical_name", "")
        if canonical:
            name_map.setdefault(canonical, []).append((mod_key, col_entry))
    elif pos == "group":
        for sub in col_entry.get("sub_columns", []):
            if sub.get("position") == "fixed":
                canonical = sub.get("canonical_name", "")
                if canonical:
                    name_map.setdefault(canonical, []).append((mod_key, sub))


def _find_parent_module(meta_module_keys):
    """Find the first metadata module key (typically module_1)."""
    return meta_module_keys[0] if meta_module_keys else None


# ---------------------------------------------------------------------------
# Pipeline entry point
# ---------------------------------------------------------------------------


@pipeline("schema-validation", timeout=3600)
@parameter(
    "structure_proposal_file",
    name="Structure Proposal JSON",
    type=File,
    required=True,
    help="The structure_proposal.json file generated by the AI analysis pipeline.",
)
@parameter(
    "excel_file",
    name="Excel Template",
    type=File,
    required=True,
    help="The original Excel template (used for example values and validation resolution).",
)
@parameter(
    "sheet_name",
    name="Sheet Name",
    type=str,
    required=False,
    help="Sheet name to process. If empty, uses the sheet specified in the proposal.",
)
def schema_validation(
    structure_proposal_file: File,
    excel_file: File,
    sheet_name: str = None,
):
    """Generate schema_validation.json from a structure proposal + Excel template."""
    current_run.log_info(f"Loading structure proposal: {structure_proposal_file}")
    proposal_path = structure_proposal_file.path
    with open(str(proposal_path), encoding="utf-8") as f:
        proposal = json.load(f)

    current_run.log_info(
        f"Proposal loaded: {len(proposal.get('sections', []))} sections, "
        f"version {proposal.get('version', '?')}"
    )

    # Load Excel
    current_run.log_info(f"Loading Excel template: {excel_file}")
    excel_path = excel_file.path
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    wb_full = openpyxl.load_workbook(str(excel_path), data_only=False)

    # Select sheet
    target_sheet = sheet_name or proposal.get("source_sheet")
    if target_sheet and target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
        ws_full = wb_full[target_sheet]
    else:
        ws = wb[wb.sheetnames[0]]
        ws_full = wb_full[wb_full.sheetnames[0]]
        current_run.log_info(f"Sheet '{target_sheet}' not found, using '{ws.title}'")

    current_run.log_info(f"Processing sheet: '{ws.title}'")

    # Extract data validations from Excel
    dv_map = extract_data_validations(ws_full)
    n_list = sum(1 for v in dv_map.values() if v.get("type") == "list")
    n_indirect = sum(1 for v in dv_map.values() if v.get("type") == "indirect")
    current_run.log_info(
        f"Data validations extracted: {n_list} list dropdowns, "
        f"{n_indirect} INDIRECT (conditional) dropdowns"
    )

    # Resolve INDIRECT dependencies into conditional_enum maps
    conditional_map = resolve_indirect_dependencies(ws_full, dv_map)
    if conditional_map:
        current_run.log_info(
            f"Conditional dropdowns resolved: "
            f"{', '.join(f'{k} depends on {v["depends_on_column"]}' for k, v in conditional_map.items())}"
        )

    wb_full.close()

    # Enrich dv_map with pre-resolved values from the proposal (fallback)
    _merge_proposal_validations(dv_map, proposal)

    # Generate schema
    current_run.log_info("Generating schema from structure proposal...")
    schema = generate_schema(proposal, ws, dv_map, conditional_map)

    # Write output
    output_path = Path(workspace.files_path) / SCHEMA_OUTPUT_FILENAME
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(schema, f, indent=2, ensure_ascii=False)

    current_run.log_info(f"Schema written to: {output_path}")

    # Summary
    n_labels = sum(
        len([k for k in v if not k.startswith("_")])
        for k, v in schema.get("label_fields", {}).items()
        if isinstance(v, dict) and not k.endswith("_columns")
    )
    n_hdr_cols = sum(
        len(schema["header_modules"][k].get("columns", []))
        for k in schema.get("header_modules", {})
        if not k.startswith("_") and k != "mismatch_message_template"
    )
    current_run.log_info(f"Schema summary: {n_labels} label fields, {n_hdr_cols} header columns")

    wb.close()


if __name__ == "__main__":
    schema_validation()
